#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#------------------------------------------------------
#--- migration_repair_doctor    v 1.0               ---
#--- MARCH 2026                                     ---
#--- ORIGINAL AUTHOR  LOCKHEED LLC   IDEO-LAB       ---
#--- GUILLAUME CLAUDE ONEILL                        ---
#--- https://www.ideo-lab.com                       ---
#------------------------------------------------------


import importlib
import json
import tempfile
import platform
import subprocess
import os
import shutil
import copy
import re

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib     import Path
from typing      import Dict, List, Optional, Sequence, Tuple
from datetime    import timezone
from datetime    import datetime


from openpyxl        import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils  import get_column_letter

from reportlab.lib           import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles    import getSampleStyleSheet
from reportlab.platypus      import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.core.management      import call_command
from django.core.management.base import CommandError
from django.apps                 import apps
from django.core.management.base import BaseCommand, CommandError
from django.db                   import connection
from django.db                   import migrations as dj_migrations
from django.db                   import models
from django.db                   import connections

from django.db.migrations.serializer        import serializer_factory
from django.db.migrations.loader            import MigrationLoader
from django.db.migrations.operations.fields import AddField
from django.db.migrations.operations.models import AddIndex, CreateModel
from django.db.migrations.migration         import Migration
from django.db.migrations.writer            import MigrationWriter
from django.db.migrations.operations.fields import AddField, AlterField
from django.db.migrations.autodetector      import MigrationAutodetector
from django.db.migrations.state             import ProjectState
from django.db.migrations.questioner        import NonInteractiveMigrationQuestioner

SAFE_INDEX_CHAR_THRESHOLD = 191


@dataclass
class ModelDbState:
    table_exists: bool
    columns: List[str] = field(default_factory=list)
    column_meta: Dict[str, Dict[str, object]] = field(default_factory=dict)
    indexes: Dict[str, List[str]] = field(default_factory=dict)
    foreign_keys: List[Dict[str, str]] = field(default_factory=list)

@dataclass
class FieldSpec:
    name: str
    field_class: str
    max_length: Optional[int] = None
    null: bool = False
    blank: bool = False
    db_index: bool = False
    primary_key: bool = False
    unique: bool = False
    related_model: Optional[str] = None
    on_delete: Optional[str] = None


@dataclass
class ModelSpec:
    model_name: str
    table_name: str
    fields: List[FieldSpec] = field(default_factory=list)


@dataclass
class AddFieldSpec:
    model_name: str
    field_name: str
    field_class: str
    related_model: Optional[str] = None
    db_column_name: Optional[str] = None
    is_relation: bool = False
    
    
@dataclass
class AlterFieldSpec:
    model_name: str
    field_name: str
    field_class: str
    db_column_name: Optional[str] = None
    is_relation: bool = False
    null: bool = False
    max_length: Optional[int] = None
    db_index: bool = False    


@dataclass
class IndexSpec:
    model_name: str
    index_name: str
    fields: List[str]
    safe: bool
    reason: str = ""




@dataclass
class OperationStatus:
    op_type: str
    model_name: str
    detail: str
    status: str
    reason: str = ""


class Command(BaseCommand):
    help = "Inspect a partially applied migration and generate repaired/followup draft migrations."

    def add_arguments(self, parser):
        parser.add_argument("--app", required=True, help="App label, for example: weglot")
        parser.add_argument(
            "--target",
            default="",
            help="Migration prefix or exact migration name, for example: 0050",
        )
        parser.add_argument(
            "--last",
            action="store_true",
            help="Inspect the latest migration on disk for the selected app.",
        )
        parser.add_argument(
            "--scan",
            action="store_true",
            help="Scan migrations and pick the most likely broken partially-applied migration.",
        )
        parser.add_argument(
            "--explore",
            action="store_true",
            help="Alias of --scan.",
        )

        parser.add_argument(
            "--write-files",
            action="store_true",
            help="Write repaired/followup migration draft files to disk",
        )
        parser.add_argument(
            "--output-dir",
            default="",
            help="Optional directory for generated files. Defaults to the app migrations folder.",
        )
        parser.add_argument(
            "--safe-threshold",
            type=int,
            default=SAFE_INDEX_CHAR_THRESHOLD,
            help="Maximum safe indexed CharField length for MySQL/MariaDB",
        )
        parser.add_argument(
            "--report-json",
            action="store_true",
            help="Also write a JSON diagnostic report",
        )
        
        parser.add_argument(
            "--export",
            choices=["excel", "pdf", "both"],
            default="",
            help="Export the diagnostic report as excel, pdf, or both.",
        )
        parser.add_argument(
            "--export-dir",
            default="",
            help="Output directory for exported files. Defaults to the system temporary directory.",
        )   
        
        parser.add_argument(
            "--open-export",
            action="store_true",
            help="Open the export directory after report generation.",
        )  
        parser.add_argument(
            "--simulation",
            action="store_true",
            help="Run in simulation mode. This is the default behavior.",
        )
        parser.add_argument(
            "--execute",
            action="store_true",
            help="Apply real changes on disk and/or in django_migrations.",
        )
        parser.add_argument(
            "--backup-dir",
            default="",
            help="Directory used to store backup files before any corrective action.",
        )
        parser.add_argument(
            "--write-repaired-migration",
            action="store_true",
            help="Write the repaired migration candidate file to disk.",
        )
        parser.add_argument(
            "--fake-record",
            action="store_true",
            help="Insert the target migration into django_migrations if the DB state is validated.",
        )
        parser.add_argument(
            "--repair-plan-json",
            action="store_true",
            help="Write a repair plan JSON file.",
        )   
        
        parser.add_argument(
            "--patch-models",
            action="store_true",
            help="Apply safe automatic fixes to models.py for dangerous indexes.",
        )
        parser.add_argument(
            "--generate-followup",
            action="store_true",
            help="Run makemigrations after model patching and repaired migration handling.",
        )
        parser.add_argument(
            "--run-makemigrations",
            action="store_true",
            help="Explicitly run makemigrations for the target app.",
        )
        parser.add_argument(
            "--run-migrate",
            action="store_true",
            help="Run migrate for the target app after followup migration generation.",
        )  
        
        parser.add_argument(
            "--clean-target-migration",
            action="store_true",
            help="Rewrite the target migration with only operations already materialized in the database.",
        )   
        
        parser.add_argument(
            "--all-in-one",
            action="store_true",
            help="Run full migration repair workflow (audit + repair + followup + migrate)",
        )  
        
        parser.add_argument(
            "--verify",
            action="store_true",
            help="Quick verification of major differences between Django models and MariaDB for the selected app.",
        )
        
        parser.add_argument(
            "--verbose",
            action="store_true",
            help="Show detailed diagnostic report.",
        )  
        
        parser.add_argument(
            "--presim",
            action="store_true",
            help="Pre-simulate the next cleaned migration expected from models.py versus MariaDB.",
        )        


    def _build_presim_summary(self, op_statuses: List[OperationStatus]) -> Dict[str, object]:
        invalid_groups = {}
        next_groups = {}
    
        for op in op_statuses:
            op_type = op.op_type
            model_name = op.model_name
            status = op.status
    
            invalid_groups.setdefault((op_type, model_name), {"count": 0, "missing": 0, "applied": 0})
    
            if status == "applied":
                invalid_groups[(op_type, model_name)]["applied"] += 1
            elif status == "missing":
                invalid_groups[(op_type, model_name)]["missing"] += 1
    
            invalid_groups[(op_type, model_name)]["count"] += 1
    
            if op_type == "AddIndex" and status == "missing":
                next_groups.setdefault(model_name, 0)
                next_groups[model_name] += 1
    
        invalid_rows = []
        for (op_type, model_name), stats in invalid_groups.items():
            if stats["applied"] == 0:
                continue
    
            if op_type == "CreateModel":
                label = "Existe déjà"
            elif op_type in {"AddField", "AlterField"}:
                label = "Existe déjà"
            elif op_type == "AddIndex":
                if stats["missing"] > 0:
                    label = f"Manque {stats['missing']} Idx"
                else:
                    label = "Existe déjà"
            else:
                label = "Existe déjà"
    
            invalid_rows.append(
                {
                    "op_type": op_type,
                    "model_name": model_name,
                    "count": stats["count"],
                    "missing": stats["missing"],
                    "label": label,
                }
            )
    
        next_rows = []
        for model_name, count in sorted(next_groups.items()):
            next_rows.append(
                {
                    "op_type": "AddIndex",
                    "model_name": model_name,
                    "count": count,
                    "label": f"Manque {count} Idx",
                }
            )
    
        return {
            "invalid_rows": invalid_rows,
            "next_rows": next_rows,
            "next_total_indexes": sum(item["count"] for item in next_rows),
        }
    
    
    def _print_presim_summary(self, app_label: str, summary: Dict[str, object]) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("IDENTIFY EACH INVALID MIGRATIONS"))
    
        for row in summary["invalid_rows"]:
            op_type = row["op_type"]
            model_name = row["model_name"]
            count = row["count"]
            label = row["label"]
    
            if op_type == "CreateModel":
                detail = f"{model_name}"
            elif op_type in {"AddField", "AlterField"}:
                detail = f"{model_name} ({count} column{'s' if count > 1 else ''})"
            elif op_type == "AddIndex":
                detail = f"{model_name} ({count} index{'es' if count > 1 else ''})"
            else:
                detail = f"{model_name} ({count})"
    
            self.stdout.write(f"- {op_type:<12} : {detail:<50} ({label})")
    
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("NEXT REMAINING MIGRATION EXPECTED"))
    
        next_rows = summary["next_rows"]
        if not next_rows:
            self.stdout.write("- none")
        else:
            for row in next_rows:
                self.stdout.write(
                    f"- AddIndex     : {row['model_name']} ({row['count']} index miss) ({row['label']})"
                )
    
        self.stdout.write("")
        self.stdout.write(f"TOTAL NEXT INDEXES: {summary['next_total_indexes']}")
        self.stdout.write("")    
        
    def _build_quick_summary(self, report: Dict[str, object]) -> Dict[str, int]:
        applied = 0
        missing = 0
    
        for item in report.get("operation_statuses", []):
            status = item.get("status", "")
            if status == "applied":
                applied += 1
            elif status == "missing":
                missing += 1
    
        return {
            "tables_checked": len(report.get("models", {})),
            "ops_applied_or_already_present": applied,
            "ops_missing_or_to_apply": missing,
            "dangerous_indexes": len(report.get("dangerous_indexes", [])),
        }
    
    
    
    def _run_verify(
        self,
        app_label: str,
        verbose: bool = False,
        use_execute: bool = False,
        ) -> None:
        
        self.stdout.write(self.style.MIGRATE_HEADING("Verify mode"))
        self.stdout.write(f"App: {app_label}")
        self.stdout.write("")
    
    
        app_config = apps.get_app_config(app_label)
        applied_index_flags = self._get_applied_model_index_flags(app_label)
        
        models_to_check = [
            model for model in app_config.get_models()
            if not model._meta.abstract and not model._meta.proxy
        ]
        total_models = len(models_to_check)
        
        table_names_to_check = [model._meta.db_table for model in models_to_check]
        prefetched_schema = self._prefetch_verify_db_schema(table_names_to_check)    
    
        errors: List[str] = []
        warnings: List[str] = []
        infos_hidden: List[str] = []
    
        pending_alter_fields: Dict[Tuple[str, str], object] = {}
        pending_alter_reasons: List[str] = []
        pending_add_indexes: Dict[Tuple[str, str], object] = {}
        pending_add_index_reasons: List[str] = []  
        pending_create_models: Dict[str, object] = {}
        pending_create_model_reasons: List[str] = []        
    
    
        if total_models == 0:
            self.stdout.write("No concrete models found.")
            self.stdout.write(self.style.SUCCESS("Status: OK"))
            return
    
        self._print_progress_bar(0, total_models, prefix="Verify")
    
        for idx, model in enumerate(models_to_check, start=1):
            meta = model._meta
            table_name = meta.db_table
            model_name = meta.model_name
    
    
    
            table_schema = prefetched_schema.get(table_name, {})
            table_exists = bool(table_schema.get("table_exists", False))
            
            if not table_exists:
                pending_create_models[meta.model_name] = model
                pending_create_model_reasons.append(
                    f"Missing table: {table_name} ({model_name})"
                )
                if self._should_refresh_progress(idx, total_models):
                    self._print_progress_bar(idx, total_models, prefix="Verify")
                continue
            
            db_columns = table_schema.get("columns", {})
            db_index_meta = table_schema.get("indexes", {})    
    
    
    
            if verbose:
                self.stdout.write(f"[OK TABLE] {table_name}")
    
            for field in meta.local_fields:
                db_column = getattr(field, "column", None) or field.name
    
                if db_column not in db_columns:
                    errors.append(f"Missing column: {table_name}.{db_column}")
                    continue
    
                if verbose:
                    self.stdout.write(f"  [OK COLUMN] {db_column}")
    
                model_max_length = getattr(field, "max_length", None)
                db_internal_size = db_columns[db_column].get("internal_size", None)    
    
    
    
    
                if model_max_length and db_internal_size:
                    try:
                        db_internal_size = int(db_internal_size)
                    except Exception:
                        db_internal_size = None
    
                    if db_internal_size is not None and int(model_max_length) != db_internal_size:
                        key = (meta.model_name, field.name)
                        pending_alter_fields[key] = copy.deepcopy(field)
                        pending_alter_reasons.append(
                            f"Length mismatch: {table_name}.{db_column} db={db_internal_size} model={model_max_length}"
                        )
    
                model_null = getattr(field, "null", False)
                db_null = db_columns[db_column].get("null_ok", None)                
                
                
                if db_null is not None and bool(db_null) != bool(model_null):
                    key = (meta.model_name, field.name)
                    pending_alter_fields[key] = copy.deepcopy(field)
                    pending_alter_reasons.append(
                        f"Null mismatch: {table_name}.{db_column} db={db_null} model={model_null}"
                    )
                    
                    
                model_unique = getattr(field, "unique", False)
                if model_unique:
                    found_unique = False
                    for _idx_name, idx_meta in db_index_meta.items():
                        if idx_meta["columns"] == [db_column] and idx_meta["unique"]:
                            found_unique = True
                            break
                    if not found_unique:
                        errors.append(f"Missing unique index: {table_name}.{db_column}")
                
                
                
                if getattr(field, "db_index", False):
                    found_db_index = False
                    for _idx_name, idx_meta in db_index_meta.items():
                        if idx_meta["columns"] == [db_column]:
                            found_db_index = True
                            break
                
                    if not found_db_index:
                        infos_hidden.append(f"Missing db_index: {table_name}.{db_column}")
                
                        applied_field_flags = (
                            applied_index_flags
                            .get(meta.model_name, {})
                            .get(field.name, {})
                        )
                        already_declared_in_applied_state = bool(
                            applied_field_flags.get("db_index", False)
                        )
                
                        if not already_declared_in_applied_state:
                            pending_add_indexes[(meta.model_name, field.name)] = copy.deepcopy(field)
                            pending_add_index_reasons.append(
                                f"New db_index declared in models: {table_name}.{db_column}"
                            )                        
    
    
    
    
            for index in getattr(meta, "indexes", []):
                expected_name = getattr(index, "name", None) or "<unnamed>"
                expected_columns = self._normalize_model_index_columns(model, index)
    
                exact_name_match = expected_name in db_index_meta
                exact_def_match = False
                exact_name_bad_def = False
                same_def_other_name = None
    
                if exact_name_match:
                    db_meta = db_index_meta[expected_name]
                    if list(db_meta["columns"]) == list(expected_columns):
                        exact_def_match = True
                    else:
                        exact_name_bad_def = True
    
                if not exact_def_match:
                    for db_name, db_meta in db_index_meta.items():
                        if list(db_meta["columns"]) == list(expected_columns):
                            same_def_other_name = db_name
                            break
    
                if exact_def_match:
                    if verbose:
                        self.stdout.write(
                            f"  [OK INDEX] {table_name}.{expected_name} ({', '.join(expected_columns)})"
                        )
                    continue
    
                if exact_name_bad_def:
                    errors.append(
                        f"Index definition mismatch: {table_name}.{expected_name} "
                        f"db={db_index_meta[expected_name]['columns']} model={expected_columns}"
                    )
                    continue
    
                if same_def_other_name:
                    warnings.append(
                        f"Index name mismatch: {table_name} expected={expected_name} "
                        f"db={same_def_other_name} cols=({', '.join(expected_columns)})"
                    )
                    continue
    
                warnings.append(
                    f"Missing composite index: {table_name}.{expected_name} "
                    f"({', '.join(expected_columns)})"
                )
    
            if self._should_refresh_progress(idx, total_models):
                self._print_progress_bar(idx, total_models, prefix="Verify")            
    
        alter_ops: List[AlterField] = []
        for (model_name, field_name), field_obj in sorted(pending_alter_fields.items()):
            alter_ops.append(
                AlterField(
                    model_name=model_name,
                    name=field_name,
                    field=copy.deepcopy(field_obj),
                    preserve_default=True,
                )
            )
    
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Verification summary"))
        self.stdout.write(f"Errors count      : {len(errors)}")
        self.stdout.write(f"Warnings count    : {len(warnings)}")
        self.stdout.write(f"Hidden infos count: {len(infos_hidden)}")
        self.stdout.write(f"Pending AlterField: {len(alter_ops)}")
        self.stdout.write(f"Pending AddIndex  : {len(pending_add_indexes)}")
        self.stdout.write(f"Pending CreateModel: {len(pending_create_models)}")
        self.stdout.write("")
    
        if warnings:
            self.stdout.write("Warnings:")
            for item in warnings:
                self.stdout.write(f"- {item}")
        else:
            self.stdout.write("Warnings: none")
    
        self.stdout.write("")
    
        if infos_hidden:
            if verbose:
                self.stdout.write("Infos hidden:")
                for item in infos_hidden:
                    self.stdout.write(f"- {item}")
            else:
                self.stdout.write(
                    f"Infos hidden: {len(infos_hidden)} item(s) hidden "
                    f"(use --verbose to display details)"
                )
        else:
            self.stdout.write("Infos hidden: none")
    
        self.stdout.write("")
    
    
        if alter_ops:
            self.stdout.write("Pending AlterField migration:")
            if verbose:
                for item in pending_alter_reasons:
                    self.stdout.write(f"- {item}")
            else:
                self.stdout.write(
                    f"- {len(alter_ops)} field(s) require AlterField migration"
                )
        else:
            self.stdout.write("Pending AlterField migration: none")
    
        self.stdout.write("")
        
        if pending_add_indexes:
            self.stdout.write("Pending AddIndex migration:")
            if verbose:
                for item in pending_add_index_reasons:
                    self.stdout.write(f"- {item}")
            else:
                self.stdout.write(
                    f"- {len(pending_add_indexes)} field(s) require AddIndex migration"
                )
        else:
            self.stdout.write("Pending AddIndex migration: none")
            
            
        self.stdout.write("")
        
        if pending_create_models:
            self.stdout.write("Pending CreateModel migration:")
            if verbose:
                for item in pending_create_model_reasons:
                    self.stdout.write(f"- {item}")
            else:
                self.stdout.write(
                    f"- {len(pending_create_models)} model(s) require CreateModel migration"
                )
        else:
            self.stdout.write("Pending CreateModel migration: none")        
    
        if errors:
            self.stdout.write(self.style.ERROR("Status: NOT OK"))
            self.stdout.write("Errors:")
            for item in errors:
                self.stdout.write(f"- {item}")
            raise CommandError("Verification failed")
    
        self.stdout.write(self.style.SUCCESS("Status: OK"))
    
    
        if use_execute:
            if pending_create_models or alter_ops or pending_add_indexes:
                self._write_followup_sync_migration(
                    app_label=app_label,
                    pending_create_models=pending_create_models,
                    alter_ops=alter_ops,
                    pending_add_indexes=pending_add_indexes,
                )
            else:
                self.stdout.write(self.style.SUCCESS("Nothing to generate."))    
    
    def _get_last_migration_name_or_none(self, app_label: str) -> Optional[str]:
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        names = sorted(
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        )
        if not names:
            return None
        return names[-1]    
    
    def _write_followup_sync_migration(
        self,
        app_label: str,
        pending_create_models: Dict[str, object],
        alter_ops: List[AlterField],
        pending_add_indexes: Dict[Tuple[str, str], object],
        ) -> Optional[Path]:
        
        if not pending_create_models and not alter_ops and not pending_add_indexes:            
            self.stdout.write(self.style.SUCCESS("No sync migration to generate."))
            return None
        
        
        latest_name = self._get_last_migration_name_or_none(app_label)
        
        if latest_name:
            latest_number = int(latest_name.split("_")[0])
            next_number = f"{latest_number + 1:04d}"
            dependencies_lines = [
                "    dependencies = [",
                f"        ('{app_label}', '{latest_name}'),",
                "    ]",
            ]
        else:
            next_number = "0001"
            dependencies_lines = [
                "    dependencies = []",
            ]
        
        migration_name = f"{next_number}_{app_label}_manual_field_sync"
        migration_path = self._get_migration_file_path(app_label, migration_name)
        
        lines: List[str] = []
        
        lines.append("from django.db import migrations, models")
        lines.append("import django.utils.timezone")
        lines.append("")
        lines.append("")        
        lines.append("class Migration(migrations.Migration):")
        lines.append("")
        lines.extend(dependencies_lines)
        lines.append("")
        lines.append("    operations = [")        
        
        
        for model_name, model_obj in sorted(pending_create_models.items()):
            create_op = dj_migrations.CreateModel(
                name=model_obj._meta.object_name,
                fields=[
                    (f.name, copy.deepcopy(f))
                    for f in model_obj._meta.local_fields
                ],
                options={},
            )
        
            rendered = self._render_python_operation(create_op)
            rendered_lines = rendered.splitlines()
            for line in rendered_lines:
                lines.append(f"        {line}")
            lines[-1] = lines[-1] + ","        
    
        # ---- AlterField ops ----
        for op in alter_ops:
            rendered = self._render_python_operation(op)
            rendered_lines = rendered.splitlines()
            for line in rendered_lines:
                lines.append(f"        {line}")
            lines[-1] = lines[-1] + ","
    
        # ---- AddIndex ops for db_index=True ----
        for (model_name, field_name), field_obj in sorted(pending_add_indexes.items()):
            db_column = getattr(field_obj, "column", None) or field_name
            index_name = f"{model_name}_{db_column}_idx"
    
            lines.append("        migrations.AddIndex(")
            lines.append(f"            model_name='{model_name}',")
            lines.append(
                f"            index=models.Index(fields=['{field_name}'], name='{index_name}'),"
            )
            lines.append("        ),")
    
        lines.append("    ]")
        lines.append("")
    
        migration_source = "\n".join(lines)
        migration_path.write_text(migration_source, encoding="utf-8")
    
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"Generated migration: {migration_path.name}"))
        self.stdout.write(f"Path: {migration_path}")
    
        return migration_path    
    
    
    
    def _print_quick_summary(self, report: Dict[str, object]) -> None:
        summary = self._build_quick_summary(report)
    
        self.stdout.write(self.style.MIGRATE_HEADING("Quick summary"))
        self.stdout.write(f"App: {report['app']}")
        self.stdout.write(f"Migration: {report['migration']}")
        self.stdout.write(f"Tables checked: {summary['tables_checked']}")
        self.stdout.write(f"Already present / applied: {summary['ops_applied_or_already_present']}")
        self.stdout.write(f"Missing / to apply: {summary['ops_missing_or_to_apply']}")
        self.stdout.write(f"Dangerous indexes: {summary['dangerous_indexes']}")
        self.stdout.write("")    
    
    
        
    def _resolve_field_db_column_name(self, field_name: str, field) -> Optional[str]:
        try:
            field_copy = copy.deepcopy(field)
    
            if hasattr(field_copy, "set_attributes_from_name"):
                field_copy.set_attributes_from_name(field_name)
    
            if hasattr(field_copy, "get_attname_column"):
                _attname, db_column = field_copy.get_attname_column()
                if db_column and str(db_column).strip():
                    return str(db_column)    
    
            remote_field = getattr(field_copy, "remote_field", None)
            if remote_field is not None:
                return f"{field_name}_id"
    
            return field_name
    
        except Exception:
            remote_field = getattr(field, "remote_field", None)
            if remote_field is not None:
                return f"{field_name}_id"
            return field_name   
        
        
    def _get_expected_add_field_column_name(self, add_field: AddFieldSpec) -> str:
        if add_field.db_column_name and str(add_field.db_column_name).strip():
            return str(add_field.db_column_name).strip()
    
        if add_field.is_relation:
            return f"{add_field.field_name}_id"
    
        return add_field.field_name    
    
        
    def _write_repaired_migration_candidate(
        self,
        app_label: str,
        migration_name: str,
        repaired_source: str,
        backup_dir: Path,
        simulate: bool,
        rewrite_in_place: bool,
        ) -> Optional[Path]:
        
        migration_path = self._get_migration_file_path(app_label, migration_name)
        candidate_path = migration_path.with_name(f"{migration_name}_repaired_candidate.py")
    
        if simulate:
            if rewrite_in_place:
                self.stdout.write(
                    self.style.WARNING(
                        f"[SIMULATION] Would backup and rewrite migration file in place: {migration_path}"
                    )
                )
            else:
                self.stdout.write(
                    self.style.WARNING(
                        f"[SIMULATION] Would write repaired migration candidate: {candidate_path}"
                    )
                )
            return None
    
        if rewrite_in_place:
            backup_path = self._backup_file(migration_path, backup_dir)
            if backup_path:
                self.stdout.write(self.style.SUCCESS(f"Backup created: {backup_path}"))
    
            migration_path.write_text(repaired_source, encoding="utf-8")
            self.stdout.write(self.style.SUCCESS(f"Rewritten migration file: {migration_path}"))
            return migration_path
    
        candidate_path.write_text(repaired_source, encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Wrote repaired migration candidate: {candidate_path}"))
        return candidate_path
    
    
    
    def _normalize_model_index_columns(self, model, index) -> List[str]:
        result = []
    
        for field_name in list(getattr(index, "fields", []) or []):
            clean_name = field_name.lstrip("-")
    
            try:
                django_field = model._meta.get_field(clean_name)
                db_column = getattr(django_field, "column", None) or clean_name
            except Exception:
                db_column = clean_name
    
            result.append(db_column)
    
        return result
    
    
        
    def _print_repair_plan(self, repair_plan: Dict[str, object]) -> None:
        self.stdout.write(self.style.MIGRATE_HEADING("Repair plan"))
        self.stdout.write(f"App: {repair_plan['app_label']}")
        self.stdout.write(f"Migration: {repair_plan['migration_name']}")
        self.stdout.write(f"Can fake record: {'YES' if repair_plan['can_fake_record'] else 'NO'}")
    
        if repair_plan["fake_record_reasons"]:
            self.stdout.write("Fake record blockers:")
            for reason in repair_plan["fake_record_reasons"]:
                self.stdout.write(f"- {reason}")
    
        self.stdout.write("Missing safe indexes:")
        if repair_plan["missing_safe_indexes"]:
            for item in repair_plan["missing_safe_indexes"]:
                self.stdout.write(
                    f"- {item['model_name']} :: {item['index_name']} ({', '.join(item['fields'])})"
                )
        else:
            self.stdout.write("- none")
    
        self.stdout.write("Skipped dangerous indexes:")
        if repair_plan["skipped_dangerous_indexes"]:
            for item in repair_plan["skipped_dangerous_indexes"]:
                self.stdout.write(
                    f"- {item['model_name']} :: {item['index_name']} ({', '.join(item['fields'])}) :: {item['reason']}"
                )
        else:
            self.stdout.write("- none")
    
        self.stdout.write("")        
        
    def _get_last_migration_name(self, app_label: str) -> str:
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        names = sorted(
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        )
        if not names:
            raise CommandError(f"No migrations found on disk for app {app_label}")
        return names[-1]    
    
    
    def _build_suggested_actions(self, op_statuses: List[OperationStatus], dangerous_indexes: List[Dict]) -> List[str]:
        actions: List[str] = []
    
        has_missing_create = any(
            op.op_type == "CreateModel" and op.status == "missing"
            for op in op_statuses
        )
        has_missing_index = any(
            op.op_type == "AddIndex" and op.status == "missing"
            for op in op_statuses
        )
    
        if dangerous_indexes:
            actions.append("Remove unsafe db_index=True flags from long text-like fields.")
            actions.append("Remove unsafe composite indexes containing long CharField or URLField values.")
    
        if has_missing_index:
            actions.append("Generate a followup migration for safe missing indexes only.")
    
        if not has_missing_create and has_missing_index:
            actions.append("Consider fake-applying the repaired migration after validating the real DB state.")
    
        if not actions:
            actions.append("No immediate corrective action detected.")
    
        return actions    
    
    
    def _scan_for_broken_candidate(self, app_label: str, safe_threshold: int) -> str:
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        disk_names = sorted(
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        )
        if not disk_names:
            raise CommandError(f"No migrations found on disk for app {app_label}")
    
        applied_names = set(self._get_applied_migration_names(app_label))
        scored_candidates: List[Dict[str, object]] = []
    
        for migration_name in disk_names:
            score = 0
            reasons: List[str] = []
    
            if migration_name not in applied_names:
                score += 10
                reasons.append("+10 unapplied")
    
            migration_module   = self._load_migration_module(app_label, migration_name)
            migration_instance = migration_module.Migration(migration_name, app_label)
            create_models      = self._collect_create_models(migration_instance, app_label)

            model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
            for key, spec in create_models.items():
                model_field_map[key] = {f.name: f for f in spec.fields}            
            
            add_indexes = self._collect_add_indexes(
                migration_instance  ,
                create_models       , 
                safe_threshold      ,
            )
    
            has_create_model = bool(create_models)
            has_add_index = bool(add_indexes)
    
            if has_create_model:
                score += 20
                reasons.append("+20 has CreateModel")
    
            if has_add_index:
                score += 20
                reasons.append("+20 has AddIndex")
    
            existing_tables: List[str] = []
            missing_index_details: List[str] = []
            dangerous_indexes: List[str] = []
    
            if create_models:
                db_state = self._inspect_db_state(list(create_models.values()))
    
                for model_key, spec in create_models.items():
                    state = db_state.get(model_key)
                    if state and state.table_exists:
                        existing_tables.append(spec.table_name)
    
                if existing_tables:
                    score += 30
                    reasons.append("+30 existing expected table(s)")
    
                for idx in add_indexes:
                    if not idx.safe:
                        dangerous_indexes.append(idx.index_name)
    
                    model_key = idx.model_name.lower()
                    state = db_state.get(model_key)
                    if not state or not state.table_exists:
                        continue
    
                    expected_fields = [
                        self._normalize_db_column_name(
                            field_name,
                            model_name=idx.model_name,
                            model_field_map=model_field_map,
                        )
                        for field_name in idx.fields
                    ]
    
                    found = False
                    for _idx_name, cols in state.indexes.items():
                        if list(cols) == list(expected_fields):
                            found = True
                            break
    
                    if not found and idx.safe:
                        missing_index_details.append(idx.index_name)    
    
                if missing_index_details:
                    score += 15
                    reasons.append("+15 missing expected index(es)")
    
                if dangerous_indexes:
                    score += 25
                    reasons.append("+25 dangerous index(es)")
    
            candidate = {
                "migration_name": migration_name,
                "score": score,
                "reasons": reasons,
                "existing_tables": existing_tables,
                "missing_indexes": missing_index_details,
                "dangerous_indexes": dangerous_indexes,
                "applied": migration_name in applied_names,
            }
            scored_candidates.append(candidate)
    
        scored_candidates.sort(
            key=lambda item: (int(item["score"]), str(item["migration_name"]))
        )
    
        best = scored_candidates[-1]
        best_score = int(best["score"])
        best_name = str(best["migration_name"])
    
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Scan mode"))
        self.stdout.write("Top candidates:")
    
        top_candidates = sorted(
            scored_candidates,
            key=lambda item: (int(item["score"]), str(item["migration_name"])),
            reverse=True,
        )[:5]
    
        for item in top_candidates:
            reasons = ", ".join(item["reasons"]) if item["reasons"] else "-"
            existing_tables = ", ".join(item["existing_tables"]) if item["existing_tables"] else "-"
            missing_indexes = ", ".join(item["missing_indexes"]) if item["missing_indexes"] else "-"
            dangerous_indexes = ", ".join(item["dangerous_indexes"]) if item["dangerous_indexes"] else "-"
            applied_flag = "applied" if item["applied"] else "unapplied"
    
            self.stdout.write(
                f"- {item['migration_name']} | score={item['score']} | {applied_flag}"
            )
            self.stdout.write(f"  reasons: {reasons}")
            self.stdout.write(f"  existing_tables: {existing_tables}")
            self.stdout.write(f"  missing_indexes: {missing_indexes}")
            self.stdout.write(f"  dangerous_indexes: {dangerous_indexes}")
    
        if best_score <= 0:
            fallback = self._get_last_migration_name(app_label)
            self.stdout.write("")
            self.stdout.write(
                self.style.WARNING(
                    f"No strong broken-migration candidate found, fallback to latest migration: {fallback}"
                )
            )
            return fallback
    
        self.stdout.write("")
        self.stdout.write(
            self.style.WARNING(
                f"Auto-detected probable broken migration: {best_name} (score={best_score})"
            )
        )
        return best_name
    
    
        
    def _resolve_target_mode(
        self,
        app_label: str,
        target: str,
        use_last: bool,
        use_scan: bool,
        safe_threshold: int,
        all_in_one: bool = False,
        ) -> str:
        
        
        selected = sum(1 for flag in [bool(target), use_last, use_scan] if flag)
        if selected > 1:
            raise CommandError("Use only one of --target, --last, or --scan/--explore.")
        
        if target:
            return self._resolve_migration_name(app_label, target)
        
        if use_last or all_in_one:
            migration_name = self._get_last_migration_name(app_label)
            self.stdout.write(self.style.WARNING(f"Using latest migration on disk: {migration_name}"))
            return migration_name
        
        if use_scan:
            return self._scan_for_broken_candidate(app_label, safe_threshold=safe_threshold)
        
        raise CommandError("You must provide one of: --target, --last, --scan/--explore.")     
    
    def _export_report_bundle(self, report: Dict, export_mode: str, export_dir: Path) -> Dict[str, str]:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        app_label = report["app"]
        migration_name = report["migration"]
        base_name = f"{app_label}_{migration_name}_{timestamp}"
    
        result: Dict[str, str] = {}
    
        if export_mode in {"excel", "both"}:
            excel_path = export_dir / f"{base_name}.xlsx"
            self._write_excel_report(report, excel_path)
            result["excel"] = str(excel_path)
    
        if export_mode in {"pdf", "both"}:
            pdf_path = export_dir / f"{base_name}.pdf"
            self._write_pdf_report(report, pdf_path)
            result["pdf"] = str(pdf_path)
    
        return result    
    
    
    def _write_excel_report(self, report: Dict, output_path: Path) -> None:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
    
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
    
        def style_header_row(ws, row_number: int, max_col: int) -> None:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row_number, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
        def autofit(ws) -> None:
            for column_cells in ws.columns:
                max_length = 0
                column_index = column_cells[0].column
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                ws.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)
    
        ws_summary["A1"] = "Migration Repair Doctor Report"
        ws_summary["A1"].font = title_font
    
        summary_rows = [
            ("AlterField count" , report["counts"].get("alter_fields", 0)),
            ("App"              , report["app"]),
            ("Migration"        , report["migration"]),
            ("Safe threshold"   , report["safe_threshold"]),
            ("CreateModel count", report["counts"]["create_models"]),
            ("AddField count"   , report["counts"]["add_fields"]),
            ("AddIndex count"   , report["counts"]["add_indexes"]),
            ("Dangerous index count", report["counts"]["dangerous_indexes"]),
            ("Generated at"     , datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
    
        start_row = 3
        for idx, (label, value) in enumerate(summary_rows, start=start_row):
            ws_summary.cell(row=idx, column=1, value=label).font = bold_font
            ws_summary.cell(row=idx, column=2, value=value)
    
        ws_summary.freeze_panes = "A3"
        autofit(ws_summary)
        
        ws_actions = wb.create_sheet("SuggestedActions")
        ws_actions.append(["Suggested Action"])
        style_header_row(ws_actions, 1, 1)
        
        for action in report.get("suggested_actions", []):
            ws_actions.append([action])
        
        ws_actions.freeze_panes = "A2"
        ws_actions.auto_filter.ref = ws_actions.dimensions
        autofit(ws_actions)        
        
    
        ws_models = wb.create_sheet("Models")
        ws_models.append(["Model", "Table", "Exists", "Columns", "Indexes", "Foreign Keys"])
        style_header_row(ws_models, 1, 6)
    
        for model_name, info in report["models"].items():
            ws_models.append([
                model_name,
                info["table_name"],
                "YES" if info["table_exists"] else "NO",
                ", ".join(info["columns"]),
                "; ".join(f"{idx}({', '.join(cols)})" for idx, cols in info["indexes"].items()),
                "; ".join(
                    f"{fk['constraint_name']}:{fk['column_name']}->{fk['referenced_table']}.{fk['referenced_column']}"
                    for fk in info["foreign_keys"]
                ),
            ])
    
        ws_models.freeze_panes = "A2"
        ws_models.auto_filter.ref = ws_models.dimensions
        autofit(ws_models)
    
        ws_ops = wb.create_sheet("Operations")
        ws_ops.append(["Operation", "Model", "Detail", "Status", "Reason"])
        style_header_row(ws_ops, 1, 5)
    
        for item in report["operation_statuses"]:
            ws_ops.append([
                item["op_type"],
                item["model_name"],
                item["detail"],
                item["status"],
                item["reason"],
            ])
    
        ws_ops.freeze_panes = "A2"
        ws_ops.auto_filter.ref = ws_ops.dimensions
        autofit(ws_ops)
    
        ws_danger = wb.create_sheet("DangerousIndexes")
        ws_danger.append(["Model", "Index Name", "Fields", "Reason"])
        style_header_row(ws_danger, 1, 4)
    
        for item in report["dangerous_indexes"]:
            ws_danger.append([
                item["model"],
                item["index_name"],
                ", ".join(item["fields"]),
                item["reason"],
            ])
    
        ws_danger.freeze_panes = "A2"
        ws_danger.auto_filter.ref = ws_danger.dimensions
        autofit(ws_danger)
        
        
        ws_timeline = wb.create_sheet("MigrationTimeline")
        ws_timeline.append([
            "Migration",
            "Applied",
            "Has CreateModel",
            "Has AddIndex",
            "Existing Table Count",
            "Missing Index Count",
            "Dangerous Index Count",
            "Inferred State",
            "Existing Tables",
            "Missing Indexes",
            "Dangerous Indexes",
        ])
        style_header_row(ws_timeline, 1, 11)
        
        for item in report.get("migration_timeline", []):
            ws_timeline.append([
                item["migration_name"],
                "YES" if item["applied"] else "NO",
                "YES" if item["has_create_model"] else "NO",
                "YES" if item["has_add_index"] else "NO",
                item["existing_table_count"],
                item["missing_index_count"],
                item["dangerous_index_count"],
                item["inferred_state"],
                ", ".join(item["existing_tables"]),
                ", ".join(item["missing_indexes"]),
                ", ".join(item["dangerous_indexes"]),
            ])
        
        ws_timeline.freeze_panes = "A2"
        ws_timeline.auto_filter.ref = ws_timeline.dimensions
        autofit(ws_timeline)        
        
    
        wb.save(output_path)    
    
    
    def _resolve_export_dir(self, export_dir: str) -> Path:
        """
        Resolve export directory depending on OS.
    
        Priority:
        1) explicit --export-dir
        2) Windows -> C:\\temp
        3) Linux   -> /tmp
        """
    
        if export_dir:
            path = Path(export_dir)
    
        else:
            system = platform.system().lower()
    
            if "windows" in system:
                path = Path("C:/temp")
    
            elif "linux" in system or "darwin" in system:
                path = Path("/tmp")
    
            else:
                # fallback safety
                path = Path(tempfile.gettempdir())
    
            path = path / "migration_repair_doctor"
    
        # Create directory safely
        try:
            path.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            self.stdout.write(
                self.style.WARNING(
                    f"Could not create export directory {path}, fallback to system temp."
                )
            )
            path = Path(tempfile.gettempdir()) / "migration_repair_doctor"
            path.mkdir(parents=True, exist_ok=True)
    
        return path    
    
    def _write_pdf_report(self, report: Dict, output_path: Path) -> None:
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
    
        styles = getSampleStyleSheet()
        story = []
    
        story.append(Paragraph("Migration Repair Doctor Report", styles["Title"]))
        story.append(Spacer(1, 12))
    
        summary_data = [
            ["App", report["app"]],
            ["Migration", report["migration"]],
            ["Safe threshold", str(report["safe_threshold"])],
            ["CreateModel count", str(report["counts"]["create_models"])],
            ["AddField count", str(report["counts"]["add_fields"])],
            ["AlterField count", str(report["counts"].get("alter_fields", 0))],
            ["AddIndex count", str(report["counts"]["add_indexes"])],            ["Dangerous index count", str(report["counts"]["dangerous_indexes"])],
            ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]
    
        story.append(Paragraph("Summary", styles["Heading2"]))
        summary_table = Table(summary_data, colWidths=[180, 500])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 16))
    
        story.append(Paragraph("Operation Status", styles["Heading2"]))
        op_data = [["Operation", "Model", "Detail", "Status", "Reason"]]
        for item in report["operation_statuses"][:40]:
            op_data.append([
                item["op_type"],
                item["model_name"],
                item["detail"],
                item["status"],
                item["reason"],
            ])
    
        op_table = Table(op_data, repeatRows=1, colWidths=[90, 120, 280, 80, 180])
        op_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(op_table)
        story.append(Spacer(1, 16))
    
        story.append(Paragraph("Dangerous Indexes", styles["Heading2"]))
        danger_data = [["Model", "Index Name", "Fields", "Reason"]]
        if report["dangerous_indexes"]:
            for item in report["dangerous_indexes"]:
                danger_data.append([
                    item["model"],
                    item["index_name"],
                    ", ".join(item["fields"]),
                    item["reason"],
                ])
        else:
            danger_data.append(["-", "-", "-", "No dangerous indexes detected"])
    
        danger_table = Table(danger_data, repeatRows=1, colWidths=[120, 180, 220, 260])
        danger_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7A1F1F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(danger_table)
        
        story.append(Spacer(1, 16))
        story.append(Paragraph("Suggested Actions", styles["Heading2"]))
        
        action_data = [["Suggested Action"]]
        for action in report.get("suggested_actions", []):
            action_data.append([action])
        
        action_table = Table(action_data, repeatRows=1, colWidths=[780])
        action_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E6B3E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(action_table)   
        
        
        story.append(Spacer(1, 16))
        story.append(Paragraph("Migration Timeline", styles["Heading2"]))
        
        timeline_data = [[
            "Migration",
            "Applied",
            "CreateModel",
            "AddIndex",
            "Existing Tbl",
            "Missing Idx",
            "Dangerous Idx",
            "State",
        ]]
        
        for item in report.get("migration_timeline", [])[:30]:
            timeline_data.append([
                item["migration_name"],
                "YES" if item["applied"] else "NO",
                "YES" if item["has_create_model"] else "NO",
                "YES" if item["has_add_index"] else "NO",
                str(item["existing_table_count"]),
                str(item["missing_index_count"]),
                str(item["dangerous_index_count"]),
                item["inferred_state"],
            ])
        
        timeline_table = Table(
            timeline_data,
            repeatRows=1,
            colWidths=[180, 55, 75, 65, 70, 70, 85, 80],
        )
        timeline_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4B4B4B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(timeline_table)        
    
        doc.build(story)    
    

    def _get_applied_migration_names(self, app_label: str) -> List[str]:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT name FROM django_migrations WHERE app = %s ORDER BY name",
                [app_label],
            )
            rows = cursor.fetchall()
        return [row[0] for row in rows]   
    
    
    
    def _can_fake_record_migration(
        self,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        db_state: Dict[str, ModelDbState],
        ) -> Tuple[bool, List[str]]:
        
        reasons: List[str] = []
    
        for model_key, spec in create_models.items():
            state = db_state.get(model_key)
        
            # CreateModel peut être légitimement manquant
            # on ne bloque pas la réparation
            if not state or not state.table_exists:
                continue    
    
    
        for add_field in add_fields:
            model_key = add_field.model_name.lower()
            state = db_state.get(model_key)
            if not state:
                reasons.append(f"Missing model state for AddField: {add_field.model_name}.{add_field.field_name}")
                continue
    
            expected_column = self._get_expected_add_field_column_name(add_field)
    
            if expected_column not in state.columns:
                reasons.append(
                    f"Missing column for AddField: {add_field.model_name}.{add_field.field_name} "
                    f"(expected DB column: {expected_column})"
                )
    
        for alter_field in alter_fields:
            model_key = alter_field.model_name.lower()
            state = db_state.get(model_key)
    
            materialized, reason = self._is_alter_field_materialized(alter_field, state)
            if not materialized:
                reasons.append(
                    f"Missing materialization for AlterField: "
                    f"{alter_field.model_name}.{alter_field.field_name} ({reason})"
                )
    
        return (len(reasons) == 0, reasons)
    
    
    
    def _write_repair_plan_json(self, repair_plan: Dict[str, object], output_path: Path) -> None:
        output_path.write_text(
            json.dumps(repair_plan, indent=2, default=str),
            encoding="utf-8",
        )    
        
        
    def _collect_missing_operations(
        self,
        migration_instance,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
        ) -> List[object]:
        
        missing_operations: List[object] = []
    
        add_field_map = {
            (item.model_name.lower(), item.field_name): item
            for item in add_fields
        }
    
        alter_field_map = {
            (item.model_name.lower(), item.field_name): item
            for item in alter_fields
        }
    
        model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
        for key, spec in create_models.items():
            model_field_map[key] = {f.name: f for f in spec.fields}
    
        for op in migration_instance.operations:
            if isinstance(op, CreateModel):
                model_key = op.name.lower()
                state = db_state.get(model_key)
                if not state or not state.table_exists:
                    missing_operations.append(op)
                continue
    
            if isinstance(op, AddField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                spec = add_field_map.get((model_key, op.name))
    
                if not state or not spec:
                    missing_operations.append(op)
                    continue
    
                expected_column = self._get_expected_add_field_column_name(spec)
                if expected_column not in state.columns:
                    missing_operations.append(op)
                continue
    
            if isinstance(op, AlterField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                spec = alter_field_map.get((model_key, op.name))
    
                if not state or not spec:
                    missing_operations.append(op)
                    continue
    
                materialized, _reason = self._is_alter_field_materialized(spec, state)
                if not materialized:
                    missing_operations.append(op)
                continue
    
            if isinstance(op, AddIndex):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
    
                expected_fields = [
                    self._normalize_db_column_name(
                        field_name,
                        model_name=op.model_name,
                        model_field_map=model_field_map,
                    )
                    for field_name in op.index.fields
                ]
    
                exists = False
                if state and state.table_exists:
                    for _idx_name, cols in state.indexes.items():
                        if list(cols) == list(expected_fields):
                            exists = True
                            break
    
                if not exists:
                    missing_operations.append(op)
                continue
    
            missing_operations.append(op)
    
        return missing_operations    
    
    
    def _build_repair_plan(
        self,
        app_label: str,
        migration_name: str,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
    ) -> Dict[str, object]:
        model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
        for key, spec in create_models.items():
            model_field_map[key] = {f.name: f for f in spec.fields}
    
        can_fake_record, fake_record_reasons = self._can_fake_record_migration(
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            db_state=db_state,
        )
    
        missing_add_fields: List[Dict[str, object]] = []
        for add_field in add_fields:
            model_key = add_field.model_name.lower()
            state = db_state.get(model_key)
            expected_column = self._get_expected_add_field_column_name(add_field)
    
            if not state or expected_column not in state.columns:
                missing_add_fields.append(
                    {
                        "model_name": add_field.model_name,
                        "field_name": add_field.field_name,
                        "db_column_name": expected_column,
                    }
                )
    
        missing_alter_fields: List[Dict[str, object]] = []
        for item in alter_fields:
            model_key = item.model_name.lower()
            state = db_state.get(model_key)
            materialized, reason = self._is_alter_field_materialized(item, state)
            if not materialized:
                missing_alter_fields.append(
                    {
                        "model_name": item.model_name,
                        "field_name": item.field_name,
                        "reason": reason,
                    }
                )
    
        missing_safe_indexes: List[Dict[str, object]] = []
        skipped_dangerous_indexes: List[Dict[str, object]] = []
    
        for idx in add_indexes:
            model_key = idx.model_name.lower()
            state = db_state.get(model_key)
            already_exists = False
    
            if state and state.table_exists:
                expected_fields = [
                    self._normalize_db_column_name(
                        field_name,
                        model_name=idx.model_name,
                        model_field_map=model_field_map,
                    )
                    for field_name in idx.fields
                ]
    
                for _idx_name, cols in state.indexes.items():
                    if cols == expected_fields:
                        already_exists = True
                        break
    
            if already_exists:
                continue
    
            payload = {
                "model_name": idx.model_name,
                "index_name": idx.index_name,
                "fields": idx.fields,
                "reason": idx.reason,
            }
    
            if idx.safe:
                missing_safe_indexes.append(payload)
            else:
                skipped_dangerous_indexes.append(payload)
    
        return {
            "app_label": app_label,
            "migration_name": migration_name,
            "can_fake_record": can_fake_record,
            "fake_record_reasons": fake_record_reasons,
            "missing_add_fields": missing_add_fields,
            "missing_alter_fields": missing_alter_fields,
            "missing_safe_indexes": missing_safe_indexes,
            "skipped_dangerous_indexes": skipped_dangerous_indexes,
        }    
    
    def _record_migration_as_applied(self, app_label: str, migration_name: str, simulate: bool) -> None:
        sql = """
            INSERT INTO django_migrations (app, name, applied)
            VALUES (%s, %s, %s)
        """
        applied_at = datetime.now()
    
        if simulate:
            self.stdout.write(
                self.style.WARNING(
                    f"[SIMULATION] Would insert django_migrations row: {app_label}.{migration_name} at {applied_at}"
                )
            )
            return
    
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM django_migrations WHERE app = %s AND name = %s",
                [app_label, migration_name],
            )
            row = cursor.fetchone()
            if row and int(row[0]) > 0:
                self.stdout.write(
                    self.style.WARNING(
                        f"Migration already recorded in django_migrations: {app_label}.{migration_name}"
                    )
                )
                return
    
            cursor.execute(sql, [app_label, migration_name, applied_at])
    
        self.stdout.write(
            self.style.SUCCESS(
                f"Recorded migration as applied: {app_label}.{migration_name}"
            )
        )    
    
    
    def _get_migration_file_path(self, app_label: str, migration_name: str) -> Path:
        app_config = apps.get_app_config(app_label)
        return Path(app_config.path) / "migrations" / f"{migration_name}.py"
    
    
    def _get_models_file_path(self, app_label: str) -> Path:
        app_config = apps.get_app_config(app_label)
        return Path(app_config.path) / "models.py"  
    
    
    def _write_followup_alter_fields_migration(
        self,
        app_label: str,
        alter_ops: List[AlterField],
        ) -> Optional[Path]:
        if not alter_ops:
            self.stdout.write(self.style.SUCCESS("No AlterField migration to generate."))
            return None
    
    
        latest_name = self._get_last_migration_name(app_label)
        latest_number = int(latest_name.split("_")[0])
        next_number = f"{latest_number + 1:04d}"
    
        migration_name = f"{next_number}_{app_label}_manual_field_sync"
        migration_path = self._get_migration_file_path(app_label, migration_name)
    
        lines: List[str] = []
        lines.append("from django.db import migrations, models")
        lines.append("")
        lines.append("")
        lines.append("class Migration(migrations.Migration):")
        lines.append("")
        lines.append("    dependencies = [")
        lines.append(f"        ('{app_label}', '{latest_name}'),")
        lines.append("    ]")
        lines.append("")
        lines.append("    operations = [")
    
        for op in alter_ops:
            rendered = self._render_python_operation(op)
            rendered_lines = rendered.splitlines()
            for line in rendered_lines:
                lines.append(f"        {line}")
            lines[-1] = lines[-1] + ","
    
        lines.append("    ]")
        lines.append("")
    
        migration_source = "\n".join(lines)
        migration_path.write_text(migration_source, encoding="utf-8")
    
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"Generated migration: {migration_path.name}"))
        self.stdout.write(f"Path: {migration_path}")
    
        return migration_path    
    
    
    def _write_cleaned_target_migration(
        self,
        app_label: str,
        migration_name: str,
        cleaned_migration: Migration,
        simulate: bool,
        backup_dir: Path,
    ) -> Optional[Path]:
        migration_path = self._get_migration_file_path(app_label, migration_name)
        writer = MigrationWriter(cleaned_migration)
        rendered = writer.as_string()
    
        if simulate:
            self.stdout.write(
                self.style.WARNING(
                    f"[SIMULATION] Would backup and rewrite cleaned target migration: {migration_path}"
                )
            )
            return None
    
        backup_path = self._backup_file(migration_path, backup_dir)
        if backup_path:
            self.stdout.write(self.style.SUCCESS(f"Backup created: {backup_path}"))
    
        migration_path.write_text(rendered, encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Rewritten cleaned migration: {migration_path}"))
        return migration_path  
    
    def _get_latest_disk_migration_name(self, app_label: str) -> Optional[str]:
        loader = MigrationLoader(None, ignore_no_migrations=True)
        names = sorted(
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        )
        return names[-1] if names else None    
    
    
    def _validate_cleaned_migration_before_fake_record(
        self,
        kept_operations: List[object],
        ) -> Tuple[bool, List[str]]:
        reasons: List[str] = []
    
        if not kept_operations:
            reasons.append("No materialized operations detected for cleaned migration.")
    
        has_create_or_addfield = any(
            isinstance(op, (CreateModel, AddField))
            for op in kept_operations
        )
        if not has_create_or_addfield:
            reasons.append("Cleaned migration contains no structural operations.")
    
        return (len(reasons) == 0, reasons)    
    
    def _build_cleaned_migration_object(
        self,
        app_label: str,
        migration_name: str,
        migration_instance,
        kept_operations: List[object],
        ) -> Migration:
        cleaned = Migration(migration_name, app_label)
        cleaned.dependencies = list(migration_instance.dependencies)
        cleaned.operations = list(kept_operations)
        cleaned.initial = getattr(migration_instance, "initial", False)
        return cleaned    
    
    def _collect_materialized_operations(
        self,
        migration_instance,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
    ) -> List[object]:
        kept_operations: List[object] = []
    
        create_model_names = set(create_models.keys())
    
        add_field_map: Dict[Tuple[str, str], AddFieldSpec] = {
            (item.model_name.lower(), item.field_name): item
            for item in add_fields
        }
    
        alter_field_map: Dict[Tuple[str, str], AlterFieldSpec] = {
            (item.model_name.lower(), item.field_name): item
            for item in alter_fields
        }
    
        index_map: Dict[Tuple[str, str], IndexSpec] = {
            (item.model_name.lower(), item.index_name): item
            for item in add_indexes
        }
    
        model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
        for key, spec in create_models.items():
            model_field_map[key] = {f.name: f for f in spec.fields}
    
        for op in migration_instance.operations:
            if isinstance(op, CreateModel):
                model_key = op.name.lower()
                if model_key not in create_model_names:
                    continue
    
                state = db_state.get(model_key)
                if state and state.table_exists:
                    kept_operations.append(op)
                continue
    
            if isinstance(op, AddField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                add_field_spec = add_field_map.get((model_key, op.name))
    
                if not state or not add_field_spec:
                    continue
    
                expected_column = self._get_expected_add_field_column_name(add_field_spec)
                if expected_column in state.columns:
                    kept_operations.append(op)
                continue
    
            if isinstance(op, AlterField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                alter_field_spec = alter_field_map.get((model_key, op.name))
    
                if not state or not alter_field_spec:
                    continue
    
                materialized, _reason = self._is_alter_field_materialized(alter_field_spec, state)
                if materialized:
                    kept_operations.append(op)
                continue
    
            if isinstance(op, AddIndex):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                idx_spec = index_map.get((model_key, op.index.name))
    
                if not state or not idx_spec:
                    continue
    
                expected_fields = [
                    self._normalize_db_column_name(
                        field_name,
                        model_name=idx_spec.model_name,
                        model_field_map=model_field_map,
                    )
                    for field_name in idx_spec.fields
                ]
    
                already_exists = False
                for _idx_name, cols in state.indexes.items():
                    if cols == expected_fields:
                        already_exists = True
                        break
    
                if already_exists:
                    kept_operations.append(op)
                continue
    
        return kept_operations    
    
    def _apply_safe_model_fixes(
        self,
        models_file_path: Path,
        patch_plan: Dict[str, object],
        simulate: bool,
        backup_dir: Path,
    ) -> Optional[Path]:
        if not models_file_path.exists():
            self.stdout.write(self.style.WARNING(f"models.py not found: {models_file_path}"))
            return None
    
        content = models_file_path.read_text(encoding="utf-8")
        original_content = content
    
        for item in patch_plan["field_fixes"]:
            field_name = item["field"]
    
            content = content.replace(
                f"{field_name} = models.CharField(max_length=512, blank=True, default=\"\", db_index=True)",
                f"{field_name} = models.CharField(max_length=512, blank=True, default=\"\")",
            )
            content = content.replace(
                f"{field_name} = models.CharField(db_index=True, max_length=512)",
                f"{field_name} = models.CharField(max_length=512)",
            )
            content = content.replace(
                f"{field_name} = models.URLField(db_index=True)",
                f"{field_name} = models.URLField()",
            )
            content = content.replace(
                f"{field_name} = models.TextField(db_index=True)",
                f"{field_name} = models.TextField()",
            )
    
        for item in patch_plan["index_fixes"]:
            index_name = item["index_name"]
            content = content.replace(
                f'models.Index(fields=["site", "canonical_path"], name="{index_name}"),',
                "",
            )
    
        if content == original_content:
            self.stdout.write(self.style.WARNING("No model patch was applied."))
            return None
    
        if simulate:
            self.stdout.write(
                self.style.WARNING(
                    f"[SIMULATION] Would patch models file: {models_file_path}"
                )
            )
            return None
    
        backup_path = self._backup_file(models_file_path, backup_dir)
        if backup_path:
            self.stdout.write(self.style.SUCCESS(f"Models backup created: {backup_path}"))
    
        models_file_path.write_text(content, encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Patched models file: {models_file_path}"))
        return models_file_path 
    
    
    
    
    
    def _run_migrate_command(self, app_label: str, simulate: bool, verbose: bool = False) -> Optional[int]:
        cmd = [
            "python",
            "manage.py",
            "migrate_safe",
            app_label,
            "--rollback-on-error",
        ]
    
        if verbose:
            cmd.append("--ddl-verbose")
    
        if simulate:
            self.stdout.write(self.style.WARNING(f"[SIMULATION] Would run: {' '.join(cmd)}"))
            return None
    
        result = subprocess.run(cmd, capture_output=True, text=True)
        stdout_text = result.stdout or ""
        stderr_text = result.stderr or ""
    
        self.stdout.write(stdout_text)
    
        run_id = None
    
        patterns = [
            r"SafeMigrationRun created:\s*id=(\d+)",
            r"SafeMigrationRun id\s*:\s*(\d+)",
        ]
    
        for pattern in patterns:
            match = re.search(pattern, stdout_text, flags=re.IGNORECASE)
            if match:
                run_id = int(match.group(1))
                break
    
        if run_id is not None:
            self.stdout.write(self.style.SUCCESS(f"Detected SafeMigrationRun id: {run_id}"))
    
        if result.returncode != 0:
            if stderr_text:
                self.stdout.write(self.style.ERROR(stderr_text))
            raise CommandError("migrate_safe failed")
    
        return run_id        
        
        
    
    def _run_makemigrations_command(self, app_label: str, simulate: bool) -> Optional[str]:
        before_name = self._get_latest_disk_migration_name(app_label)
        cmd = ["python", "manage.py", "makemigrations", app_label]
    
        if simulate:
            self.stdout.write(self.style.WARNING(f"[SIMULATION] Would run: {' '.join(cmd)}"))
            return None
    
        result = subprocess.run(cmd, capture_output=True, text=True)
        self.stdout.write(result.stdout)
        if result.returncode != 0:
            self.stdout.write(self.style.ERROR(result.stderr))
            raise CommandError("makemigrations failed")
    
        after_name = self._get_latest_disk_migration_name(app_label)
        if after_name and after_name != before_name:
            self.stdout.write(self.style.SUCCESS(f"Generated followup migration: {after_name}"))
            return after_name
    
        self.stdout.write(self.style.WARNING("No new migration was generated by makemigrations."))
        return None
    
    def _resolve_model_table_name(self, app_label: str, model_name: str) -> str:
        app_config = apps.get_app_config(app_label)
    
        for model in app_config.get_models():
            if model._meta.model_name.lower() == model_name.lower():
                return model._meta.db_table
    
        return f"{app_label}_{model_name.lower()}"    
    
    
    def _ensure_missing_tables(self, app_label, create_models, db_state):
        
        for model_key, spec in create_models.items():
            state = db_state.get(model_key)
    
            if state and state.table_exists:
                continue
    
            model = None
            self.stdout.write(
                f"DEBUG ensure_missing_tables -> app={app_label} "
                f"model_name={spec.model_name} table={spec.table_name}"
            )            
    
            try:
                model = apps.get_model(app_label, spec.model_name)
            except LookupError:
                app_config = apps.get_app_config(app_label)
    
                for candidate in app_config.get_models():
                    if candidate._meta.model_name.lower() == model_key.lower():
                        model = candidate
                        break
    
            if model is None:
                raise LookupError(
                    f"Unable to resolve Django model for "
                    f"{app_label}.{spec.model_name}"
                )
    
            with connection.schema_editor() as schema_editor:
                schema_editor.create_model(model)
    
            self.stdout.write(
                self.style.WARNING(f"Table created: {spec.table_name}")
            )            
            
    def _write_followup_file(self, app_label, migration_name, code):
    
        migrations_module = importlib.import_module(f"{app_label}.migrations")
        migrations_dir = Path(migrations_module.__file__).parent
    
        next_number = self._next_migration_number(migration_name)
    
        filename = f"{next_number}_auto_followup.py"
    
        path = migrations_dir / filename
    
        with open(path, "w", encoding="utf-8") as f:
            f.write(code)
    
        self.stdout.write(
            self.style.SUCCESS(f"Follow-up written: {filename}")
        ) 
        
    
    def _rewrite_migration_file(
        self,
        app_label,
        migration_name,
        migration_instance,
        operations,
        ):
        
        module = migration_instance.__module__
        migration_module = importlib.import_module(module)
    
        code = self._render_clean_migration(
            migration_module,
            migration_instance,
            operations,
        )
    
        path = migration_module.__file__
    
        with open(path, "w", encoding="utf-8") as f:
            f.write(code)
    
        self.stdout.write(self.style.SUCCESS("Migration rewritten"))    
    
    def _render_clean_migration(
        self,
        migration_module,
        migration_instance,
        operations,
        ) -> str:
        
        app_label = migration_instance.app_label
        migration_name = migration_instance.name
    
        header = self._render_header(migration_module)
        deps_source = self._render_dependencies(
            list(getattr(migration_instance, "dependencies", []))
        )
    
        operations_lines = [
            self._indent(self._render_python_operation(op), 2)
            for op in operations
        ]
    
        operations_block = ",\n".join(operations_lines)
    
        if operations_block:
            operations_block = "\n" + operations_block + "\n"
        else:
            operations_block = "\n"
    
        return (
            f"{header}\n\n"
            "class Migration(migrations.Migration):\n\n"
            f"    dependencies = {deps_source}\n\n"
            "    operations = ["
            f"{operations_block}"
            "    ]\n"
        )    
    
    
    
    
    
    def _run_all_in_one(self, app_label: str, simulate: bool, verbose: bool = False) -> None:
        self.stdout.write(self.style.SUCCESS("Running ALL-IN-ONE migration repair"))
    
        generated_name = self._run_makemigrations_command(
            app_label=app_label,
            simulate=simulate,
        )
    
        if simulate:
            self.stdout.write(self.style.WARNING("[SIMULATION] all-in-one stops after makemigrations simulation"))
            return
    
        if not generated_name:
            self.stdout.write(self.style.WARNING("No new migration generated. Nothing to clean."))
            return
    
        migration_name = generated_name
        migration_module = self._load_migration_module(app_label, migration_name)
        migration_instance = migration_module.Migration(migration_name, app_label)
    
        create_models = self._collect_create_models(migration_instance, app_label)
        add_fields = self._collect_add_fields(migration_instance)
        alter_fields = self._collect_alter_fields(migration_instance)
        add_indexes = self._collect_add_indexes(
            migration_instance,
            create_models,
            SAFE_INDEX_CHAR_THRESHOLD,
        )
    
        touched_models = self._collect_touched_models(
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
        )
    
        db_state = self._inspect_db_state_for_models(
            app_label=app_label,
            model_names=touched_models,
            create_models=create_models,
        )
    
        kept_operations = self._collect_missing_operations(
            migration_instance=migration_instance,
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
            db_state=db_state,
        )
    
        skipped_operations = []
        for op in migration_instance.operations:
            if op not in kept_operations:
                skipped_operations.append(op.__class__.__name__)
    
        self._rewrite_migration_file(
            app_label,
            migration_name,
            migration_instance,
            kept_operations,
        )
    
        self.stdout.write(self.style.SUCCESS(f"Migration cleaned in place: {migration_name}"))
    
        call_command("migrate", app_label)
    
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("All-in-one summary"))
        self.stdout.write(f"Generated migration: {migration_name}")
        self.stdout.write(f"Kept operations: {len(kept_operations)}")
        self.stdout.write(f"Skipped operations: {len(skipped_operations)}")
    
        if verbose and skipped_operations:
            self.stdout.write("Skipped details:")
            for item in skipped_operations:
                self.stdout.write(f"- {item}")    
    
    
    
    
    
    
    def _backup_file(self, source_path: Path, backup_dir: Path) -> Optional[Path]:
        if not source_path.exists() or not source_path.is_file():
            return None
    
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{source_path.name}.{timestamp}.bak"
        target_path = backup_dir / backup_name
        shutil.copy2(source_path, target_path)
        return target_path    
    
    
    def _resolve_backup_dir(self, backup_dir: str) -> Path:
        if backup_dir:
            path = Path(backup_dir)
        else:
            system = platform.system().lower()
    
            if "windows" in system:
                path = Path("C:/temp") / "migration_repair_doctor_backups"
            elif "linux" in system or "darwin" in system:
                path = Path("/tmp") / "migration_repair_doctor_backups"
            else:
                path = Path(tempfile.gettempdir()) / "migration_repair_doctor_backups"
    
        path.mkdir(parents=True, exist_ok=True)
        return path    
    
    
    def _get_table_index_meta(self, table_name: str) -> Dict[str, Dict[str, object]]:
        result: Dict[str, Dict[str, object]] = {}
    
        with connection.cursor() as cursor:
            introspected = connection.introspection.get_constraints(cursor, table_name)
    
        for name, meta in introspected.items():
            if not meta.get("index") and not meta.get("unique"):
                continue
    
            result[name] = {
                "columns": list(meta.get("columns") or []),
                "unique": bool(meta.get("unique", False)),
                "primary_key": bool(meta.get("primary_key", False)),
            }
    
        return result    
    
    
    def _is_alter_field_materialized(self, alter_field, state):
    
        if not state or not state.table_exists:
            return False, "table missing"
    
        column = alter_field.db_column_name or alter_field.field_name
    
        if column not in state.columns:
            return False, f"column missing: {column}"
    
        meta = state.column_meta.get(column, {})
    
        # ---- VARCHAR LENGTH ----
        if alter_field.max_length:
            db_len = meta.get("internal_size")
    
            if db_len and int(db_len) != int(alter_field.max_length):
                return False, f"length mismatch db={db_len} model={alter_field.max_length}"
    
        # ---- NULLABLE ----
        if alter_field.null is not None:
            db_null = meta.get("null_ok")
    
            if db_null is not None and bool(db_null) != bool(alter_field.null):
                return False, f"null mismatch db={db_null} model={alter_field.null}"
    
        # ---- DEFAULT ----
        if alter_field.default is not None:
            db_default = meta.get("default")
    
            if db_default != alter_field.default:
                return False, f"default mismatch db={db_default} model={alter_field.default}"
    
        # ---- UNIQUE ----
        if alter_field.unique:
            found_unique = False
    
            for idx, cols in state.indexes.items():
                if cols == [column]:
                    found_unique = True
                    break
    
            if not found_unique:
                return False, "missing unique index"
    
        return True, ""    
    
    
    def _build_migration_timeline(self, app_label: str, safe_threshold: int) -> List[Dict[str, object]]:
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        disk_names = sorted(
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        )
    
        applied_names = set(self._get_applied_migration_names(app_label))
        timeline: List[Dict[str, object]] = []
    
        for migration_name in disk_names:
            migration_module = self._load_migration_module(app_label, migration_name)
            migration_instance = migration_module.Migration(migration_name, app_label)
    
            create_models = self._collect_create_models(migration_instance, app_label)
            model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
            for key, spec in create_models.items():
                model_field_map[key] = {f.name: f for f in spec.fields}            
            
            add_indexes = self._collect_add_indexes(
                migration_instance,
                create_models,
                safe_threshold,
            )
    
            existing_tables: List[str] = []
            missing_indexes: List[str] = []
            dangerous_indexes: List[str] = []
    
            if create_models:
                db_state = self._inspect_db_state(list(create_models.values()))
    
                for model_key, spec in create_models.items():
                    state = db_state.get(model_key)
                    if state and state.table_exists:
                        existing_tables.append(spec.table_name)
    
                for idx in add_indexes:
                    if not idx.safe:
                        dangerous_indexes.append(idx.index_name)
    
                    model_key = idx.model_name.lower()
                    state = db_state.get(model_key)
                    if not state or not state.table_exists:
                        continue
    
                    expected_fields = [
                        self._normalize_db_column_name(
                            field_name,
                            model_name=idx.model_name,
                            model_field_map=model_field_map,
                        )
                        for field_name in idx.fields
                    ]
                    
                    found = False
                    for _idx_name, cols in state.indexes.items():
                        if cols == expected_fields:
                            found = True
                            break
    
                    if not found:
                        missing_indexes.append(idx.index_name)
    
            is_applied = migration_name in applied_names
            has_create_model = bool(create_models)
            has_add_index = bool(add_indexes)
    
            inferred_state = "pending"
            if is_applied:
                inferred_state = "applied"
            elif existing_tables and missing_indexes:
                inferred_state = "partial"
            elif existing_tables:
                inferred_state = "partial"
            elif not is_applied:
                inferred_state = "pending"
    
            timeline.append(
                {
                    "migration_name": migration_name,
                    "applied": is_applied,
                    "has_create_model": has_create_model,
                    "has_add_index": has_add_index,
                    "existing_table_count": len(existing_tables),
                    "missing_index_count": len(missing_indexes),
                    "dangerous_index_count": len(dangerous_indexes),
                    "existing_tables": existing_tables,
                    "missing_indexes": missing_indexes,
                    "dangerous_indexes": dangerous_indexes,
                    "inferred_state": inferred_state,
                }
            )
    
        return timeline    
    
    
    def _print_model_patch_plan(self, patch_plan: Dict[str, object]) -> None:
        self.stdout.write(self.style.MIGRATE_HEADING("Model patch plan"))
        self.stdout.write(f"App: {patch_plan['app_label']}")
    
        self.stdout.write("Field fixes:")
        if patch_plan["field_fixes"]:
            for item in patch_plan["field_fixes"]:
                self.stdout.write(
                    f"- {item['model']}.{item['field']} :: {item['reason']}"
                )
        else:
            self.stdout.write("- none")
    
        self.stdout.write("Index fixes:")
        if patch_plan["index_fixes"]:
            for item in patch_plan["index_fixes"]:
                self.stdout.write(
                    f"- {item['model']} :: {item['index_name']} ({item['fields']}) :: {item['reason']}"
                )
        else:
            self.stdout.write("- none")
    
        self.stdout.write("")    
    
    
    def _build_model_patch_plan(
        self,
        app_label: str,
        safe_threshold: int,
        ) -> Dict[str, object]:
        app_config = apps.get_app_config(app_label)
    
        field_fixes: List[Dict[str, str]] = []
        index_fixes: List[Dict[str, str]] = []
    
        for model in app_config.get_models():
            meta = model._meta
            if meta.abstract or meta.proxy:
                continue
    
            field_map = {field.name: field for field in meta.local_fields}
    
            for field in meta.local_fields:
                max_length = getattr(field, "max_length", None)
    
                if field.__class__.__name__ in {"CharField", "SlugField", "URLField"}:
                    if getattr(field, "db_index", False) and max_length and max_length > safe_threshold:
                        field_fixes.append(
                            {
                                "model": meta.object_name,
                                "field": field.name,
                                "reason": f"{field.__class__.__name__}({max_length}) with db_index=True exceeds threshold",
                            }
                        )
    
                if field.__class__.__name__ == "TextField" and getattr(field, "db_index", False):
                    field_fixes.append(
                        {
                            "model": meta.object_name,
                            "field": field.name,
                            "reason": "TextField with db_index=True is unsafe",
                        }
                    )
    
            for index in getattr(meta, "indexes", []):
                index_name = getattr(index, "name", "<unnamed_index>")
                fields = list(getattr(index, "fields", []) or [])
    
                dangerous = False
                for field_name in fields:
                    clean_name = field_name.lstrip("-")
                    field = field_map.get(clean_name)
                    if field is None:
                        continue
    
                    field_type = field.__class__.__name__
                    max_length = getattr(field, "max_length", None)
    
                    if field_type == "TextField":
                        dangerous = True
                        break
    
                    if field_type in {"CharField", "SlugField", "URLField"}:
                        if max_length and max_length > safe_threshold:
                            dangerous = True
                            break
    
                if dangerous:
                    index_fixes.append(
                        {
                            "model": meta.object_name,
                            "index_name": index_name,
                            "fields": ", ".join(fields),
                            "reason": "Composite index contains unsafe long/text field",
                        }
                    )
    
        return {
            "app_label": app_label,
            "field_fixes": field_fixes,
            "index_fixes": index_fixes,
        }    
    
    
    
    def _open_export_directory(self, export_dir: Path) -> None:
        system = platform.system().lower()
    
        try:
            if "windows" in system:
                os.startfile(str(export_dir))
                self.stdout.write(self.style.SUCCESS(f"Opened export directory: {export_dir}"))
    
            elif "linux" in system:
                subprocess.Popen(["xdg-open", str(export_dir)])
                self.stdout.write(self.style.SUCCESS(f"Opened export directory: {export_dir}"))
    
            elif "darwin" in system:
                subprocess.Popen(["open", str(export_dir)])
                self.stdout.write(self.style.SUCCESS(f"Opened export directory: {export_dir}"))
    
            else:
                self.stdout.write(
                    self.style.WARNING(
                        f"Automatic folder opening is not supported on this platform: {system}"
                    )
                )
    
        except Exception as exc:
            self.stdout.write(
                self.style.WARNING(
                    f"Could not open export directory {export_dir}: {exc}"
                )
            )    


    def handle(self, *args, **options):
        app_label          = options["app"]
        target             = (options["target"] or "").strip()
        use_last           = options["last"]
        use_scan           = options["scan"] or options["explore"]
        export_mode        = (options["export"] or "").strip()
        export_dir         = (options["export_dir"] or "").strip()  
        open_export        = options["open_export"]
        use_simulation     = options["simulation"]
        use_execute        = options["execute"]
        backup_dir         = (options["backup_dir"] or "").strip()
        write_repaired_migration = options["write_repaired_migration"]
        fake_record        = options["fake_record"]
        repair_plan_json   = options["repair_plan_json"]   
        patch_models       = options["patch_models"]
        generate_followup  = options["generate_followup"]
        run_makemigrations = options["run_makemigrations"]
        run_migrate        = options["run_migrate"]        
        write_files        = options["write_files"]
        output_dir         = options["output_dir"].strip()
        safe_threshold     = options["safe_threshold"]
        report_json        = options["report_json"]
        all_in_one = options.get("all_in_one", False)
        clean_target_migration = options["clean_target_migration"]   
        app_label          = options["app"]
        verify             = options["verify"]
        verbose            = options["verbose"]   
        presim             = options["presim"]
        
        if use_simulation and use_execute:
            raise CommandError("Use either --simulation or --execute, not both.")
        
        if not use_simulation and not use_execute:
            use_simulation = True        

        app_config = apps.get_app_config(app_label)
        
        if verify:
            self._run_verify(
                app_label,
                verbose=verbose,
                use_execute=use_execute,
            )
            return        
        
        
        
        if presim:
            app_config = apps.get_app_config(app_label)
        
            missing_by_model = {}
            total_missing = 0
        
            for model in app_config.get_models():
                meta = model._meta
        
                if meta.abstract or meta.proxy:
                    continue
        
                table_name = meta.db_table
                db_indexes = self._get_table_indexes(table_name)
        
                missing_indexes = []
        
                for index in getattr(meta, "indexes", []):
                    expected_cols = []
        
                    for field_name in list(getattr(index, "fields", []) or []):
                        clean_name = field_name.lstrip("-")
        
                        try:
                            django_field = meta.get_field(clean_name)
                            db_column = getattr(django_field, "column", None) or clean_name
                        except Exception:
                            db_column = clean_name
        
                        expected_cols.append(db_column)
        
                    found = False
                    for _idx_name, cols in db_indexes.items():
                        if list(cols) == list(expected_cols):
                            found = True
                            break
        
                    if not found:
                        missing_indexes.append(
                            {
                                "name": getattr(index, "name", "<unnamed>"),
                                "fields": expected_cols,
                            }
                        )
        
                if missing_indexes:
                    missing_by_model[meta.model_name] = missing_indexes
                    total_missing += len(missing_indexes)
        
            self.stdout.write("")
            self.stdout.write("NEXT REMAINING MIGRATION EXPECTED")
        
            if not missing_by_model:
                self.stdout.write("- none")
            else:
                for model_name, items in sorted(missing_by_model.items()):
                    self.stdout.write(
                        f"- AddIndex     : {model_name} ({len(items)} index miss)"
                    )
        
                    if verbose:
                        for item in items:
                            self.stdout.write(
                                f"    - {item['name']} ({', '.join(item['fields'])})"
                            )
        
            self.stdout.write("")
            self.stdout.write(f"TOTAL NEXT INDEXES: {total_missing}")
            self.stdout.write("")
            
            if use_execute:
                self._generate_total_migration_from_models(app_label)
        
            return        
        
        
        
        
        migration_name = self._resolve_target_mode(
            app_label=app_label,
            target=target,
            use_last=use_last,
            use_scan=use_scan,
            safe_threshold=safe_threshold,
            all_in_one=all_in_one,
        )
        
        resolved_backup_dir = self._resolve_backup_dir(backup_dir)
        migration_file_path = self._get_migration_file_path(app_label, migration_name)
        models_file_path    = self._get_models_file_path(app_label)        
        
        migration_module    = self._load_migration_module(app_label, migration_name)
        migration_instance  = migration_module.Migration(migration_name, app_label)

        create_models = self._collect_create_models(migration_instance, app_label)
        add_fields    = self._collect_add_fields(migration_instance)
        alter_fields  = self._collect_alter_fields(migration_instance)
        add_indexes   = self._collect_add_indexes(migration_instance, create_models, safe_threshold)
        
        touched_models = self._collect_touched_models(
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
        )
        
        db_state = self._inspect_db_state_for_models(
            app_label=app_label,
            model_names=touched_models,
            create_models=create_models,
        )
        
        
        if all_in_one:
            self._run_all_in_one(
                app_label=app_label,
                simulate=use_simulation,
                verbose=verbose,
            )
            return        
        
        
        
        op_statuses = self._build_operation_statuses(
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
            db_state=db_state,
        )
        
        report = self._build_report(
            app_label=app_label,
            migration_name=migration_name,
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
            db_state=db_state,
            op_statuses=op_statuses,
            safe_threshold=safe_threshold,
        )        
        
        repair_plan = self._build_repair_plan(
            app_label=app_label,
            migration_name=migration_name,
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
            db_state=db_state,
        )        
        
        model_patch_plan = self._build_model_patch_plan(
            app_label=app_label,
            safe_threshold=safe_threshold,
        )        

        if verbose:
            self._print_report(report)
            self._print_repair_plan(repair_plan)
            self._print_model_patch_plan(model_patch_plan)
        else:
            self._print_quick_summary(report)        
        
        
        if export_mode:
            resolved_export_dir = self._resolve_export_dir(export_dir)

            export_paths = self._export_report_bundle(
                report=report,
                export_mode=export_mode,
                export_dir=resolved_export_dir,
            )

            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING("Exported report files"))
            self.stdout.write(f"Export directory: {resolved_export_dir}")
            for label, path in export_paths.items():
                self.stdout.write(f"- {label}: {path}")   
            if open_export:
                self._open_export_directory(resolved_export_dir)            
                


        kept_operations = self._collect_missing_operations(
            migration_instance=migration_instance,
            create_models=create_models,
            add_fields=add_fields,
            alter_fields=alter_fields,
            add_indexes=add_indexes,
            db_state=db_state,
        )
        
        cleaned_migration = self._build_cleaned_migration_object(
            app_label=app_label,
            migration_name=migration_name,
            migration_instance=migration_instance,
            kept_operations=kept_operations,
        )        
        
    
        can_write_cleaned, cleaned_reasons = self._validate_cleaned_migration_before_fake_record(
                kept_operations=kept_operations,
            )        
        
        
        if clean_target_migration or fake_record or repair_plan_json:
            if use_simulation:
                self.stdout.write(self.style.WARNING("Running in SIMULATION mode"))
            else:
                self.stdout.write(self.style.WARNING("Running in EXECUTE mode"))

            self.stdout.write(f"Backup directory: {resolved_backup_dir}")

            if repair_plan_json:
                repair_plan_path = resolved_backup_dir / f"{migration_name}_repair_plan.json"
                if use_simulation:
                    self.stdout.write(
                        self.style.WARNING(
                            f"[SIMULATION] Would write repair plan JSON: {repair_plan_path}"
                        )
                    )
                else:
                    self._write_repair_plan_json(repair_plan, repair_plan_path)
                    self.stdout.write(self.style.SUCCESS(f"Wrote repair plan JSON: {repair_plan_path}"))

            if clean_target_migration:
                if not can_write_cleaned:
                    self.stdout.write(
                        self.style.WARNING(
                            "Clean target migration blocked because validation failed."
                        )
                    )
                    for reason in cleaned_reasons:
                        self.stdout.write(f"- {reason}")
                else:
                    self._write_cleaned_target_migration(
                        app_label=app_label,
                        migration_name=migration_name,
                        cleaned_migration=cleaned_migration,
                        simulate=use_simulation,
                        backup_dir=resolved_backup_dir,
                    )

            if fake_record:
                if not repair_plan["can_fake_record"]:
                    self.stdout.write(
                        self.style.WARNING(
                            "Fake record blocked because DB state validation failed."
                        )
                    )
                elif not clean_target_migration:
                    self.stdout.write(
                        self.style.WARNING(
                            "Fake record blocked: rerun with --clean-target-migration first."
                        )
                    )
                elif not can_write_cleaned:
                    self.stdout.write(
                        self.style.WARNING(
                            "Fake record blocked because cleaned migration validation failed."
                        )
                    )
                else:
                    self._record_migration_as_applied(
                        app_label=app_label,
                        migration_name=migration_name,
                        simulate=use_simulation,
                    )





        if patch_models:
            self._apply_safe_model_fixes(
                        models_file_path=models_file_path,
                        patch_plan=model_patch_plan,
                        simulate=use_simulation,
                        backup_dir=resolved_backup_dir,
                    )
    
        if generate_followup or run_makemigrations:
            self._run_makemigrations_command(
                        app_label=app_label,
                        simulate=use_simulation,
                    )
    
        if run_migrate:
            self._run_migrate_command(
                        app_label=app_label,
                        simulate=use_simulation,
                    )        
        

        if report_json:
            base_output_dir = (
                Path(output_dir)
                if output_dir
                else Path(app_config.path) / "migrations"
            )
            base_output_dir.mkdir(parents=True, exist_ok=True)

            report_path = base_output_dir / f"{migration_name}_repair_report.json"
            report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
            self.stdout.write(self.style.SUCCESS(f"Wrote {report_path}"))


    def _resolve_migration_name(self, app_label: str, target: str) -> str:
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        names = [
            name for (app, name) in loader.disk_migrations.keys()
            if app == app_label
        ]
        exact = [name for name in names if name == target]
        if exact:
            return exact[0]

        prefix = [name for name in names if name.startswith(target)]
        if len(prefix) == 1:
            return prefix[0]
        if not prefix:
            raise CommandError(f"No migration found for {app_label}.{target}")
        raise CommandError(
            f"Multiple migrations match prefix {target}: {', '.join(sorted(prefix))}"
        )

    def _load_migration_module(self, app_label: str, migration_name: str):
        module_path = f"{app_label}.migrations.{migration_name}"
        try:
            return importlib.import_module(module_path)
        except Exception as exc:
            raise CommandError(f"Could not import migration module {module_path}: {exc}") from exc

    def _collect_create_models(self, migration_instance, app_label: str) -> Dict[str, ModelSpec]:
        result: Dict[str, ModelSpec] = {}

        for op in migration_instance.operations:
            if not isinstance(op, CreateModel):
                continue

            model_name = op.name
            table_name = self._get_create_model_table_name(app_label, op)
            field_specs: List[FieldSpec] = []

            for field_name, field in op.fields:
                field_specs.append(self._field_to_spec(field_name, field))

            result[model_name.lower()] = ModelSpec(
                model_name=model_name,
                table_name=table_name,
                fields=field_specs,
            )

        return result
    
    
    
    
    def _inspect_db_state_for_models(
        self,
        app_label: str,
        model_names: Sequence[str],
        create_models: Optional[Dict[str, ModelSpec]] = None,
        ) -> Dict[str, ModelDbState]:
        
        all_tables = set(connection.introspection.table_names())
        result: Dict[str, ModelDbState] = {}
        create_models = create_models or {}
    
        for model_name in model_names:
            model_key = model_name.lower()
    
            if model_key in create_models:
                table_name = create_models[model_key].table_name
            else:
                table_name = self._resolve_model_table_name(app_label, model_name)
    
            table_exists = table_name in all_tables
            
            
            columns = []
            column_meta = {}
            indexes = {}
            foreign_keys = []
            
            if table_exists:
                with connection.cursor() as cursor:
                    description = connection.introspection.get_table_description(cursor, table_name)
            
                for col in description:
                    columns.append(col.name)
            
                    column_meta[col.name] = {
                        "type_code": getattr(col, "type_code", None),
                        "internal_size": getattr(col, "internal_size", None),
                        "null_ok": getattr(col, "null_ok", None),
                        "default": getattr(col, "default", None),
                    }
            
                indexes = self._get_table_indexes(table_name)
                foreign_keys = self._get_foreign_keys(table_name)
            
            result[model_key] = ModelDbState(
                table_exists=table_exists,
                columns=columns,
                column_meta=column_meta,
                indexes=indexes,
                foreign_keys=foreign_keys,
            )
            
        return result    
    
    
    
    
    def _collect_touched_models(
        self,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        ) -> List[str]:
        names = set()
    
        for key in create_models.keys():
            names.add(key.lower())
    
        for item in add_fields:
            names.add(item.model_name.lower())
    
        for item in alter_fields:
            names.add(item.model_name.lower())
    
        for item in add_indexes:
            names.add(item.model_name.lower())
    
        return sorted(names)
    
    

    def _collect_alter_fields(self, migration_instance) -> List[AlterFieldSpec]:
        result: List[AlterFieldSpec] = []
    
        for op in migration_instance.operations:
            if not isinstance(op, AlterField):
                continue
    
            remote_field = getattr(op.field, "remote_field", None)
            is_relation = remote_field is not None
            db_column_name = self._resolve_field_db_column_name(op.name, op.field)
    
            result.append(
                AlterFieldSpec(
                    model_name=op.model_name,
                    field_name=op.name,
                    field_class=op.field.__class__.__name__,
                    db_column_name=db_column_name,
                    is_relation=is_relation,
                    null=getattr(op.field, "null", False),
                    max_length=getattr(op.field, "max_length", None),
                    db_index=getattr(op.field, "db_index", False),
                )
            )
    
        return result    


    def _collect_add_fields(self, migration_instance) -> List[AddFieldSpec]:
        result: List[AddFieldSpec] = []

        for op in migration_instance.operations:
            if not isinstance(op, AddField):
                continue
            
            remote_field = getattr(op.field, "remote_field", None)
            is_relation = remote_field is not None
            db_column_name = self._resolve_field_db_column_name(op.name, op.field)                
            
            result.append(
                AddFieldSpec(
                    model_name=op.model_name,
                    field_name=op.name,
                    field_class=op.field.__class__.__name__,
                    related_model=getattr(op.field.remote_field, "model", None)
                    if hasattr(op.field, "remote_field")
                    else None,
                    db_column_name=db_column_name,
                    is_relation=is_relation,
                )
            )            
            
        return result
    
    

    def _collect_add_indexes(
        self,
        migration_instance,
        create_models: Dict[str, ModelSpec],
        safe_threshold: int,
    ) -> List[IndexSpec]:
        result: List[IndexSpec] = []

        model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
        for key, spec in create_models.items():
            model_field_map[key] = {f.name: f for f in spec.fields}

        for op in migration_instance.operations:
            if not isinstance(op, AddIndex):
                continue

            model_name = op.model_name
            fields = list(getattr(op.index, "fields", []) or [])
            safe, reason = self._is_safe_index(
                model_name=model_name,
                fields=fields,
                model_field_map=model_field_map,
                safe_threshold=safe_threshold,
            )

            result.append(
                IndexSpec(
                    model_name=model_name,
                    index_name=op.index.name,
                    fields=fields,
                    safe=safe,
                    reason=reason,
                )
            )

        return result

    def _field_to_spec(self, field_name: str, field) -> FieldSpec:
        related_model = None
        on_delete = None

        remote_field = getattr(field, "remote_field", None)
        if remote_field is not None:
            related_model = str(getattr(remote_field, "model", "")) or None
            on_delete_fn = getattr(remote_field, "on_delete", None)
            if on_delete_fn is not None:
                on_delete = getattr(on_delete_fn, "__name__", str(on_delete_fn))

        return FieldSpec(
            name=field_name,
            field_class=field.__class__.__name__,
            max_length=getattr(field, "max_length", None),
            null=getattr(field, "null", False),
            blank=getattr(field, "blank", False),
            db_index=getattr(field, "db_index", False),
            primary_key=getattr(field, "primary_key", False),
            unique=getattr(field, "unique", False),
            related_model=related_model,
            on_delete=on_delete,
        )

    def _get_create_model_table_name(self, app_label: str, op: CreateModel) -> str:
        options = getattr(op, "options", {}) or {}
        db_table = options.get("db_table")
        if db_table:
            return db_table
        return f"{app_label}_{op.name.lower()}"

    def _inspect_db_state(self, model_specs: Sequence[ModelSpec]) -> Dict[str, ModelDbState]:
        all_tables = set(connection.introspection.table_names())
        result: Dict[str, ModelDbState] = {}

        for spec in model_specs:
            table_exists = spec.table_name in all_tables
            columns: List[str] = []
            indexes: Dict[str, List[str]] = {}
            foreign_keys: List[Dict[str, str]] = []

            if table_exists:
                description = connection.introspection.get_table_description(connection.cursor(), spec.table_name)
                columns = [col.name for col in description]
                indexes = self._get_table_indexes(spec.table_name)
                foreign_keys = self._get_foreign_keys(spec.table_name)

            result[spec.model_name.lower()] = ModelDbState(
                table_exists=table_exists,
                columns=columns,
                indexes=indexes,
                foreign_keys=foreign_keys,
            )

        return result

    def _get_table_indexes(self, table_name: str) -> Dict[str, List[str]]:
        vendor = connection.vendor
        if vendor == "mysql":
            return self._get_mysql_indexes(table_name)

        introspected = connection.introspection.get_constraints(connection.cursor(), table_name)
        result: Dict[str, List[str]] = {}
        for name, data in introspected.items():
            if data.get("index") or data.get("unique"):
                result[name] = list(data.get("columns", []) or [])
        return result

    def _get_mysql_indexes(self, table_name: str) -> Dict[str, List[str]]:
        sql = """
        SELECT INDEX_NAME, SEQ_IN_INDEX, COLUMN_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        ORDER BY INDEX_NAME, SEQ_IN_INDEX
        """
        rows: List[Tuple[str, int, str]]
        with connection.cursor() as cursor:
            cursor.execute(sql, [table_name])
            rows = cursor.fetchall()

        result: Dict[str, List[str]] = {}
        for index_name, _seq, column_name in rows:
            result.setdefault(index_name, []).append(column_name)
        return result

    def _get_foreign_keys(self, table_name: str) -> List[Dict[str, str]]:
        vendor = connection.vendor
        if vendor != "mysql":
            return []

        sql = """
        SELECT
            CONSTRAINT_NAME,
            COLUMN_NAME,
            REFERENCED_TABLE_NAME,
            REFERENCED_COLUMN_NAME
        FROM information_schema.KEY_COLUMN_USAGE
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
          AND REFERENCED_TABLE_NAME IS NOT NULL
        ORDER BY CONSTRAINT_NAME, ORDINAL_POSITION
        """
        with connection.cursor() as cursor:
            cursor.execute(sql, [table_name])
            rows = cursor.fetchall()

        result = []
        for constraint_name, column_name, ref_table, ref_column in rows:
            result.append(
                {
                    "constraint_name": constraint_name,
                    "column_name": column_name,
                    "referenced_table": ref_table,
                    "referenced_column": ref_column,
                }
            )
        return result
    
    
    def _generate_total_migration_from_models(self, app_label: str) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("GENERATE TOTAL MIGRATION"))
    
        connection_obj = connections["default"]
        loader = MigrationLoader(connection_obj, ignore_no_migrations=True)
    
        from_state = loader.project_state()
        to_state = ProjectState.from_apps(apps)
    
        autodetector = MigrationAutodetector(
            from_state=from_state,
            to_state=to_state,
            questioner=NonInteractiveMigrationQuestioner(
                specified_apps={app_label},
                dry_run=False,
            ),
        )
    
        changes = autodetector.changes(
            graph=loader.graph,
            trim_to_apps={app_label},
            convert_apps={app_label},
        )
    
        app_changes = changes.get(app_label, [])
        if not app_changes:
            self.stdout.write(self.style.SUCCESS("No migration to generate."))
            return
    
        written = 0
    
        for migration_obj in app_changes:
            writer = MigrationWriter(migration_obj)
            migration_path = writer.path
            migration_source = writer.as_string()
    
            Path(migration_path).write_text(migration_source, encoding="utf-8")
    
            written += 1
            self.stdout.write(self.style.SUCCESS(f"Generated migration: {Path(migration_path).name}"))
            self.stdout.write(f"Path: {migration_path}")
    
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"Total generated migration files: {written}"))




    def _build_operation_statuses(
        self,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
    ) -> List[OperationStatus]:
        result: List[OperationStatus] = []
    
        model_field_map: Dict[str, Dict[str, FieldSpec]] = {}
        for key, spec in create_models.items():
            model_field_map[key] = {f.name: f for f in spec.fields}
    
        for model_key, spec in create_models.items():
            state = db_state.get(model_key)
            exists = bool(state and state.table_exists)
    
            result.append(
                OperationStatus(
                    op_type="CreateModel",
                    model_name=spec.model_name,
                    detail=spec.table_name,
                    status="applied" if exists else "missing",
                    reason="" if exists else "table missing",
                )
            )
    
        for add_field in add_fields:
            model_key = add_field.model_name.lower()
            state = db_state.get(model_key)
    
            if not state or not state.table_exists:
                result.append(
                    OperationStatus(
                        op_type="AddField",
                        model_name=add_field.model_name,
                        detail=f"{add_field.field_name} [db:{self._get_expected_add_field_column_name(add_field)}]",
                        status="missing",
                        reason="table missing",
                    )
                )
                continue
    
            expected_column = self._get_expected_add_field_column_name(add_field)
            exists = expected_column in state.columns
    
            result.append(
                OperationStatus(
                    op_type="AddField",
                    model_name=add_field.model_name,
                    detail=f"{add_field.field_name} [db:{expected_column}]",
                    status="applied" if exists else "missing",
                    reason="" if exists else "column missing",
                )
            )
    
        for alter_field in alter_fields:
            model_key = alter_field.model_name.lower()
            state = db_state.get(model_key)
    
            materialized, reason = self._is_alter_field_materialized(alter_field, state)
    
            result.append(
                OperationStatus(
                    op_type="AlterField",
                    model_name=alter_field.model_name,
                    detail=f"{alter_field.field_name} [db:{alter_field.db_column_name or alter_field.field_name}]",
                    status="applied" if materialized else "missing",
                    reason=reason,
                )
            )
    
        for idx in add_indexes:
            model_key = idx.model_name.lower()
            state = db_state.get(model_key)
    
            if not state or not state.table_exists:
                result.append(
                    OperationStatus(
                        op_type="AddIndex",
                        model_name=idx.model_name,
                        detail=f"{idx.index_name} ({', '.join(idx.fields)})",
                        status="missing",
                        reason="table missing",
                    )
                )
                continue
    
            expected_fields = [
                self._normalize_db_column_name(
                    field_name,
                    model_name=idx.model_name,
                    model_field_map=model_field_map,
                )
                for field_name in idx.fields
            ]
    
            already_exists = False
            for _idx_name, cols in state.indexes.items():
                if cols == expected_fields:
                    already_exists = True
                    break
    
            reason = ""
            if not already_exists and not idx.safe:
                reason = idx.reason or "unsafe index"
    
            result.append(
                OperationStatus(
                    op_type="AddIndex",
                    model_name=idx.model_name,
                    detail=f"{idx.index_name} ({', '.join(idx.fields)})",
                    status="applied" if already_exists else "missing",
                    reason=reason,
                )
            )
    
        return result    
    
    def _serialize_index(self, index) -> str:
        try:
            path, args, kwargs = index.deconstruct()
        except Exception as exc:
            raise ValueError(f"Could not deconstruct index {index!r}: {exc}") from exc
    
        args_str = ", ".join(repr(arg) for arg in args)
        kwargs_str = ", ".join(f"{key}={value!r}" for key, value in kwargs.items())
    
        joined = ", ".join(x for x in [args_str, kwargs_str] if x)
        return f"{path}({joined})"
    
    
    def _serialize_field(self, field) -> str:
        try:
            _name, path, args, kwargs = field.deconstruct()
        except Exception as exc:
            raise ValueError(f"Could not deconstruct field {field!r}: {exc}") from exc
    
        if path.startswith("django.db.models."):
            path = "models." + path.split(".")[-1]
    
        def _serialize_value(value):
            try:
                text, _imports = serializer_factory(value).serialize()
                return text
            except Exception:
                return repr(value)
    
        args_str = ", ".join(_serialize_value(arg) for arg in args)
        kwargs_str = ", ".join(
            f"{key}={_serialize_value(value)}"
            for key, value in kwargs.items()
        )
    
        joined = ", ".join(x for x in [args_str, kwargs_str] if x)
        return f"{path}({joined})"    
    
    
    
    def _render_python_operation(self, op) -> str:
        if isinstance(op, CreateModel):
            field_lines = []
            for name, field in op.fields:
                field_repr = self._serialize_field(field)
                field_lines.append(f"        ({name!r}, {field_repr}),")
    
            options_repr = ""
            if getattr(op, "options", None):
                options_repr = f",\n    options={op.options!r}"
    
    
    
            bases_repr = ""
            if getattr(op, "bases", None):
                base_names = [getattr(base, "__name__", str(base)) for base in op.bases]
                if base_names != ["Model"]:
                    bases_repr = f",\n    bases={op.bases!r}"    
    
    
            managers_repr = ""
            if getattr(op, "managers", None):
                managers_repr = f",\n    managers={op.managers!r}"
    
            fields_block = "\n".join(field_lines)
            if fields_block:
                fields_block = "\n" + fields_block + "\n"
            else:
                fields_block = "\n"
    
            return (
                "migrations.CreateModel(\n"
                f"    name={op.name!r},\n"
                "    fields=["
                f"{fields_block}"
                "    ]"
                f"{options_repr}"
                f"{bases_repr}"
                f"{managers_repr},\n"
                ")"
            )
    
        if isinstance(op, AddField):
            return "\n".join([
                "migrations.AddField(",
                f"    model_name={op.model_name!r},",
                f"    name={op.name!r},",
                f"    field={self._serialize_field(op.field)},",
                f"    preserve_default={getattr(op, 'preserve_default', True)!r},",
                ")",
            ])
    
        if isinstance(op, AlterField):
            return "\n".join([
                "migrations.AlterField(",
                f"    model_name={op.model_name!r},",
                f"    name={op.name!r},",
                f"    field={self._serialize_field(op.field)},",
                f"    preserve_default={getattr(op, 'preserve_default', True)!r},",
                ")",
            ])
    
        if isinstance(op, AddIndex):
            return "\n".join([
                "migrations.AddIndex(",
                f"    model_name={op.model_name!r},",
                f"    index={self._serialize_index(op.index)},",
                ")",
            ])
    
        raise ValueError(f"Unsupported operation type: {op.__class__.__name__}")


    def _get_latest_applied_migration_name(self, app_label: str) -> Optional[str]:
        names = self._get_applied_migration_names(app_label)
        if not names:
            return None
        return sorted(names)[-1]
    
    
    def _get_applied_model_index_flags(
        self,
        app_label: str,
        ) -> Dict[str, Dict[str, Dict[str, object]]]:
        """
        Return the field-level index flags from the latest APPLIED migration state.
    
        Result format:
        {
            "translationunit": {
                "engine": {
                    "db_index": False,
                    "unique": False,
                    "column": "engine",
                },
                ...
            },
            ...
        }
        """
        latest_applied = self._get_latest_applied_migration_name(app_label)
        if not latest_applied:
            return {}
    
        loader = MigrationLoader(connection, ignore_no_migrations=True)
        state = loader.project_state(nodes=[(app_label, latest_applied)])
        state_apps = state.apps
    
        result: Dict[str, Dict[str, Dict[str, object]]] = {}
    
        try:
            app_config = state_apps.get_app_config(app_label)
        except LookupError:
            return {}
    
        for model in app_config.get_models():
            meta = model._meta
            if meta.abstract or meta.proxy:
                continue
    
            field_map: Dict[str, Dict[str, object]] = {}
            for field in meta.local_fields:
                field_map[field.name] = {
                    "db_index": bool(getattr(field, "db_index", False)),
                    "unique": bool(getattr(field, "unique", False)),
                    "column": getattr(field, "column", None) or field.name,
                }
    
            result[meta.model_name] = field_map
    
        return result    


    def _prefetch_verify_db_schema(self, table_names: List[str]) -> Dict[str, Dict[str, object]]:
        """
        Prefetch DB schema for verification in ONE pass as much as possible.
    
        Returned format:
        {
            "table_name": {
                "columns": {
                    "col_name": {
                        "internal_size": 40,
                        "null_ok": True,
                        "default": None,
                    },
                    ...
                },
                "indexes": {
                    "idx_name": {
                        "columns": ["col1", "col2"],
                        "unique": False,
                        "primary_key": False,
                    }
                }
            }
        }
        """
        result: Dict[str, Dict[str, object]] = {}
        if not table_names:
            return result
    
        all_tables = set(connection.introspection.table_names())
        existing_tables = [t for t in table_names if t in all_tables]
    
        for table_name in table_names:
            result[table_name] = {
                "table_exists": table_name in all_tables,
                "columns": {},
                "indexes": {},
            }
    
        if not existing_tables:
            return result
    
        vendor = connection.vendor
    
        # Fast path for MariaDB / MySQL
        if vendor == "mysql":
            placeholders = ", ".join(["%s"] * len(existing_tables))
    
            sql_columns = f"""
                SELECT
                    TABLE_NAME,
                    COLUMN_NAME,
                    CHARACTER_MAXIMUM_LENGTH,
                    IS_NULLABLE,
                    COLUMN_DEFAULT
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME IN ({placeholders})
                ORDER BY TABLE_NAME, ORDINAL_POSITION
            """
    
            sql_indexes = f"""
                SELECT
                    s.TABLE_NAME,
                    s.INDEX_NAME,
                    s.SEQ_IN_INDEX,
                    s.COLUMN_NAME,
                    CASE WHEN s.NON_UNIQUE = 0 THEN 1 ELSE 0 END AS IS_UNIQUE,
                    CASE WHEN s.INDEX_NAME = 'PRIMARY' THEN 1 ELSE 0 END AS IS_PRIMARY
                FROM information_schema.STATISTICS s
                WHERE s.TABLE_SCHEMA = DATABASE()
                  AND s.TABLE_NAME IN ({placeholders})
                ORDER BY s.TABLE_NAME, s.INDEX_NAME, s.SEQ_IN_INDEX
            """
    
            with connection.cursor() as cursor:
                cursor.execute(sql_columns, existing_tables)
                for table_name, column_name, char_len, is_nullable, column_default in cursor.fetchall():
                    result[table_name]["columns"][column_name] = {
                        "internal_size": char_len,
                        "null_ok": str(is_nullable).upper() == "YES",
                        "default": column_default,
                    }
    
                cursor.execute(sql_indexes, existing_tables)
                for table_name, index_name, seq_in_index, column_name, is_unique, is_primary in cursor.fetchall():
                    idx_map = result[table_name]["indexes"]
                    if index_name not in idx_map:
                        idx_map[index_name] = {
                            "columns": [],
                            "unique": bool(is_unique),
                            "primary_key": bool(is_primary),
                        }
                    idx_map[index_name]["columns"].append(column_name)
    
            return result
    
        # Generic fallback
        for table_name in existing_tables:
            with connection.cursor() as cursor:
                description = connection.introspection.get_table_description(cursor, table_name)
                constraints = connection.introspection.get_constraints(cursor, table_name)
    
            for col in description:
                result[table_name]["columns"][col.name] = {
                    "internal_size": getattr(col, "internal_size", None),
                    "null_ok": getattr(col, "null_ok", None),
                    "default": getattr(col, "default", None),
                }
    
            for name, meta in constraints.items():
                if not meta.get("index") and not meta.get("unique"):
                    continue
                result[table_name]["indexes"][name] = {
                    "columns": list(meta.get("columns") or []),
                    "unique": bool(meta.get("unique", False)),
                    "primary_key": bool(meta.get("primary_key", False)),
                }
    
        return result    


    def _build_report(
        self,
        app_label: str,
        migration_name: str,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
        op_statuses: List[OperationStatus],
        safe_threshold: int,
    ) -> Dict[str, object]:
        
        dangerous_indexes = [
            {
                "model"      : idx.model_name ,
                "index_name" : idx.index_name ,
                "fields"     : idx.fields     ,
                "reason"     : idx.reason     ,
            }
            for idx in add_indexes
            if not idx.safe
        ]

        applied_create = [op.detail for op in op_statuses if op.op_type == "CreateModel" and op.status == "applied"]
        missing_indexes = [op.detail for op in op_statuses if op.op_type == "AddIndex" and op.status == "missing"]

        return {
            "app"               : app_label       ,
            "migration"         : migration_name  ,
            "safe_threshold"    : safe_threshold  ,
            "suggested_actions" : self._build_suggested_actions(
                op_statuses       = op_statuses   ,
                dangerous_indexes = dangerous_indexes,
            ),  
            
            "migration_timeline": self._build_migration_timeline(
                app_label=app_label,
                safe_threshold=safe_threshold,
            ),  
            
            "models": {
                model_name: {
                    "table_name"   : spec.table_name,
                    "table_exists" : db_state[model_name].table_exists,
                    "columns"      : db_state[model_name].columns,
                    "indexes"      : db_state[model_name].indexes,
                    "foreign_keys" : db_state[model_name].foreign_keys,
                }
                for model_name, spec in create_models.items()
            },
            
            "counts": {
                "create_models"     : len(create_models),
                "add_fields"        : len(add_fields),
                "alter_fields"      : len(alter_fields),
                "add_indexes"       : len(add_indexes),
                "dangerous_indexes" : len(dangerous_indexes),
            },
            
            "summary": {
                "applied_create_models": applied_create,
                "missing_indexes"      : missing_indexes,
            },
            
            "dangerous_indexes"    : dangerous_indexes,
            "operation_statuses"   : [
                {
                    "op_type"    : op.op_type,
                    "model_name" : op.model_name,
                    "detail"     : op.detail,
                    "status"     : op.status,
                    "reason"     : op.reason,
                }
                for op in op_statuses
            ],
        }
    
    
    def _print_progress_bar(self, current: int, total: int, prefix: str = "Progress", width: int = 34) -> None:
        total = max(total, 1)
        current = max(0, min(current, total))
        percent = int((current / total) * 100)
    
        filled = int(width * current / total)
        bar = "#" * filled + "-" * (width - filled)
    
        end_char = "\n" if current >= total else "\r"
        self.stdout.write(
            f"{prefix} [{bar}] {percent:3d}% ({current}/{total})",
            ending=end_char,
        )
    
    
    def _print_verify_summary(
        self,
        errors: List[str],
        warnings: List[str],
        infos_hidden: List[str],
        verbose: bool = False,
        ) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Verification summary"))
    
        self.stdout.write(f"Errors count      : {len(errors)}")
        self.stdout.write(f"Warnings count    : {len(warnings)}")
        self.stdout.write(f"Hidden infos count: {len(infos_hidden)}")
    
        self.stdout.write("")
    
        if warnings:
            self.stdout.write("Warnings:")
            for item in warnings:
                self.stdout.write(f"- {item}")
        else:
            self.stdout.write("Warnings: none")
    
        self.stdout.write("")
    
        if infos_hidden:
            if verbose:
                self.stdout.write("Infos hidden:")
                for item in infos_hidden:
                    self.stdout.write(f"- {item}")
            else:
                self.stdout.write(
                    f"Infos hidden: {len(infos_hidden)} item(s) hidden "
                    f"(use --verbose to display details)"
                )
        else:
            self.stdout.write("Infos hidden: none")
    
        self.stdout.write("")
    
        if errors:
            self.stdout.write(self.style.ERROR("Status: NOT OK"))
            self.stdout.write("Errors:")
            for item in errors:
                self.stdout.write(f"- {item}")
            raise CommandError("Verification failed")
    
        self.stdout.write(self.style.SUCCESS("Status: OK"))    

    def _print_report(self, report: Dict) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Migration Repair Doctor"))
        self.stdout.write(f"App: {report['app']}")
        self.stdout.write(f"Migration: {report['migration']}")
        self.stdout.write("")

        self.stdout.write(self.style.MIGRATE_LABEL("Model / table state"))
        for model_key, info in report["models"].items():
            state = "EXISTS" if info["table_exists"] else "MISSING"
            self.stdout.write(f"- {model_key} -> {info['table_name']} [{state}]")
            if info["table_exists"]:
                self.stdout.write(f"  columns: {', '.join(info['columns'])}")
                index_parts = []
                for idx_name, cols in info["indexes"].items():
                    index_parts.append(f"{idx_name}({', '.join(cols)})")
                self.stdout.write(f"  indexes: {', '.join(index_parts) if index_parts else '-'}")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_LABEL("Operation status"))
        for op in report["operation_statuses"]:
            suffix = f" | {op['reason']}" if op["reason"] else ""
            self.stdout.write(
                f"- {op['op_type']} {op['model_name']} :: {op['detail']} -> {op['status']}{suffix}"
            )

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_LABEL("Dangerous indexes"))
        if report["dangerous_indexes"]:
            for item in report["dangerous_indexes"]:
                self.stdout.write(
                    f"- {item['model']} :: {item['index_name']} ({', '.join(item['fields'])}) :: {item['reason']}"
                )
        else:
            self.stdout.write("- none")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_LABEL("Migration timeline"))
        for item in report.get("migration_timeline", [])[-10:]:
            self.stdout.write(
                f"- {item['migration_name']} :: "
                f"applied={'yes' if item['applied'] else 'no'} :: "
                f"state={item['inferred_state']} :: "
                f"existing_tables={item['existing_table_count']} :: "
                f"missing_indexes={item['missing_index_count']} :: "
                f"dangerous_indexes={item['dangerous_index_count']}"
            )

        self.stdout.write("")        
        

    def _render_repaired_migration(
        self,
        app_label: str,
        migration_name: str,
        migration_module,
        migration_instance,
        create_models: Dict[str, ModelSpec],
        add_fields: List[AddFieldSpec],
        db_state: Dict[str, ModelDbState],
        ) -> str:
        
        header = self._render_header(migration_module)
        repaired_class_name = "Migration"

        deps_source = self._render_dependencies(migration_instance.dependencies)
        operations_lines: List[str] = []

        for op in migration_instance.operations:
            if isinstance(op, CreateModel):
                model_key = op.name.lower()
                
                if db_state.get(model_key) and db_state[model_key].table_exists:
                    operations_lines.append(self._indent(self._render_python_operation(op), 2))                    
                    
                    
            elif isinstance(op, AddField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
            
                expected_column = self._resolve_field_db_column_name(op.name, op.field)
            
                if state and expected_column in state.columns:
                    operations_lines.append(self._indent(self._render_python_operation(op), 2))

        operations_block = ",\n".join(operations_lines)
        if not operations_block:
            operations_block = ""
        
        return (
            f"{header}\n\n"
            "class Migration(migrations.Migration):\n\n"
            f"    dependencies = {deps_source}\n\n"
            "    operations = [\n"
            f"{operations_block}\n"
            "    ]\n"
        )        

    def _should_refresh_progress(self, current: int, total: int) -> bool:
        if current <= 1 or current >= total:
            return True
        if total <= 25:
            return True
        step = max(1, total // 20)  # ~5% updates
        return current % step == 0
    
    

    def _render_followup_migration(
        self,
        app_label: str,
        migration_name: str,
        migration_module,
        migration_instance,
        add_fields: List[AddFieldSpec],
        alter_fields: List[AlterFieldSpec],
        add_indexes: List[IndexSpec],
        db_state: Dict[str, ModelDbState],
        ) -> str:
        
        header = self._render_header(migration_module)
        deps_source = self._render_dependencies([(app_label, migration_name)])
    
        add_field_map: Dict[Tuple[str, str], AddFieldSpec] = {
            (item.model_name.lower(), item.field_name): item
            for item in add_fields
        }
    
        alter_field_map: Dict[Tuple[str, str], AlterFieldSpec] = {
            (item.model_name.lower(), item.field_name): item
            for item in alter_fields
        }
    
        index_map: Dict[Tuple[str, str], IndexSpec] = {
            (item.model_name.lower(), item.index_name): item
            for item in add_indexes
        }
    
        def _resolve_index_db_columns(model_name: str, field_names: List[str]) -> List[str]:
            app_config = apps.get_app_config(app_label)
            target_model = None
    
            for model in app_config.get_models():
                if model._meta.model_name.lower() == model_name.lower():
                    target_model = model
                    break
    
            resolved: List[str] = []
    
            for field_name in field_names:
                db_col = field_name
    
                if target_model is not None:
                    try:
                        django_field = target_model._meta.get_field(field_name)
                        db_col = getattr(django_field, "column", None) or field_name
                    except Exception:
                        db_col = field_name
    
                resolved.append(db_col)
    
            return resolved
    
        operations_lines: List[str] = []
    
        for op in migration_instance.operations:
            if isinstance(op, AddField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                spec = add_field_map.get((model_key, op.name))
    
                if not state or not spec:
                    continue
    
                expected_column = self._get_expected_add_field_column_name(spec)
    
                if expected_column not in state.columns:
                    operations_lines.append(
                        self._indent(self._render_python_operation(op), 2)
                    )
                continue
    
            if isinstance(op, AlterField):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                spec = alter_field_map.get((model_key, op.name))
    
                if not state or not spec:
                    continue
    
                materialized, _reason = self._is_alter_field_materialized(spec, state)
                if not materialized:
                    operations_lines.append(
                        self._indent(self._render_python_operation(op), 2)
                    )
                continue
    
            if isinstance(op, AddIndex):
                model_key = op.model_name.lower()
                state = db_state.get(model_key)
                spec = index_map.get((model_key, op.index.name))
    
                if not state or not spec:
                    continue
    
                if not spec.safe:
                    continue
    
                expected_fields = _resolve_index_db_columns(
                    model_name=op.model_name,
                    field_names=list(op.index.fields),
                )
    
                already_exists = False
                if state.table_exists:
                    for _idx_name, cols in state.indexes.items():
                        if list(cols) == list(expected_fields):
                            already_exists = True
                            break
    
                if not already_exists:
                    operations_lines.append(
                        self._indent(self._render_python_operation(op), 2)
                    )
                continue
    
        operations_block = ",\n".join(operations_lines)
        if operations_block:
            operations_block = "\n" + operations_block + "\n"
        else:
            operations_block = "\n"
            
        return (
            f"{header}\n\n"
            "class Migration(migrations.Migration):\n\n"
            f"    dependencies = {deps_source}\n\n"
            "    operations = ["
            f"{operations_block}"
            "    ]\n"
        )


    def _render_header(self, migration_module) -> str:
        lines = [
            "from django.conf import settings",
            "from django.db import migrations, models",
            "import django.db.models.deletion",
            "import django.utils.timezone",
        ]
        if getattr(migration_module, "weglot", None):
            lines.append("import weglot.models")
        return "\n".join(lines)

    def _render_dependencies(self, dependencies) -> str:
        rendered = []
        for dep in dependencies:
            if isinstance(dep, tuple):
                rendered.append(repr(dep))
            else:
                rendered.append(repr(dep))
        joined = ",\n        ".join(rendered)
        return f"[\n        {joined},\n    ]"

    def _next_migration_number(self, migration_name: str) -> str:
        prefix = migration_name.split("_", 1)[0]
        if not prefix.isdigit():
            return "9999"
        return str(int(prefix) + 1).zfill(len(prefix))
    
    

    def _next_migration_filename(self, app_label: str, migration_name: str) -> str:
        next_number = self._next_migration_number(migration_name)
        return f"{next_number}_benchmark_followup_operations.py"    
    


    def _normalize_db_column_name(
        self,
        field_name: str,
        model_name: Optional[str] = None,
        model_field_map: Optional[Dict[str, Dict[str, FieldSpec]]] = None,
        ) -> str:
        
        if not field_name:
            return field_name
        
        clean_name = str(field_name).lstrip("-")    
        if clean_name.endswith("_id"):
            return clean_name
    
        if model_name and model_field_map:
            field_specs = model_field_map.get(model_name.lower(), {})
            field_spec = field_specs.get(clean_name)
            if field_spec and field_spec.related_model:
                return f"{clean_name}_id"
    
        return clean_name    
    

    def _is_safe_index(
        self,
        model_name: str,
        fields: Sequence[str],
        model_field_map: Dict[str, Dict[str, FieldSpec]],
        safe_threshold: int,
    ) -> Tuple[bool, str]:
        field_specs = model_field_map.get(model_name.lower(), {})
        for field_name in fields:
            clean_name = field_name.lstrip("-")
            field_spec = field_specs.get(clean_name)
            if field_spec is None:
                continue

            if field_spec.field_class == "TextField":
                return False, f"TextField in index: {clean_name}"

            if field_spec.field_class in {"CharField", "SlugField", "URLField"}:
                if field_spec.max_length and field_spec.max_length > safe_threshold:
                    return (
                        False,
                        f"{field_spec.field_class}({field_spec.max_length}) exceeds safe threshold on {clean_name}",
                    )

        return True, ""

    def _indent(self, text: str, level: int) -> str:
        pad = " " * (level * 4)
        return "\n".join(f"{pad}{line}" for line in text.splitlines())